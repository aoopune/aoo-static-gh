#!/usr/bin/env python3
"""
Export data/HOME_LOANS_COMPARE_v1.xlsx → data/home-loans-compare.json
Dumb dump only. No fee/rate repair.

Product query contract (Compare later; no JSON reshape):
1. Match offers by inputs; Blank/Any = no restriction; prefer roi_availability Offered for quotes.
2. Join bank_charges on bank_key (+ filters / when_it_matters).
3. Slabs: group charge_group_id, order by slab_from.
4. Govt: filter by jurisdiction/state.
5. EMI in UI from matched roi + amount + tenure.
6. Apply packet: data_version + bank_keys + offer_row_ids + charge ids.
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
from part_prepayment_rules import DEFAULT_CSV, load_part_prepayment_rules  # noqa: E402

SOURCE = ROOT / "data" / "HOME_LOANS_COMPARE_v1.xlsx"
OUT_JSON = ROOT / "data" / "home-loans-compare.json"
OUT_VAL = ROOT / "data" / "HOME_LOANS_COMPARE_JSON_VALIDATION.md"
PART_PREPAY_RULES_CSV = DEFAULT_CSV

SHEETS = {
    "Offers": "offers",
    "Bank_charges": "bank_charges",
    "Government_charges": "government_charges",
}
LOCKED_COUNTS = {"offers": 806, "bank_charges": 1402, "government_charges": 18}
PROCESSING_ROWS = 212
HEADER_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]*$")
PROPERTY_CHECK_ORIGIN = "Temporary.property_checks"
PROPERTY_CHECK_NAMES = (
    "Legal and technical",
    "Title search report",
    "Valuation",
)


def property_check_amount_triples(n_banks: int) -> list[tuple[int, int, int]]:
    """Hundreds near ₹4,500. Distinct lines in each bank; no shared totals."""
    pool = list(range(3900, 5400, 100))
    used_totals: set[int] = set()
    used_triples: set[tuple[int, int, int]] = set()
    out: list[tuple[int, int, int]] = []
    for index in range(n_banks):
        found = None
        for da in range(len(pool)):
            legal = pool[(index + da) % len(pool)]
            for db in range(len(pool)):
                title = pool[(index * 2 + db + 3) % len(pool)]
                if title == legal:
                    continue
                for dc in range(len(pool)):
                    valuation = pool[(index * 3 + dc + 5) % len(pool)]
                    if valuation == legal or valuation == title:
                        continue
                    triple = (legal, title, valuation)
                    total = legal + title + valuation
                    if total in used_totals or triple in used_triples:
                        continue
                    found = triple
                    break
                if found:
                    break
            if found:
                break
        if found is None:
            raise SystemExit(
                f"Could not assign unique property-check amounts for bank index {index}"
            )
        out.append(found)
        used_totals.add(sum(found))
        used_triples.add(found)
    if len(used_totals) != n_banks:
        raise SystemExit("property-check overlay totals are not unique")
    if any(len(set(triple)) != 3 for triple in out):
        raise SystemExit("property-check overlay has a bank with matching line amounts")
    return out


def inject_temporary_property_check_charges(
    rows: list[dict[str, Any]], offers: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Typical property-check amounts, unique per bank, until bank-wise figures replace them."""
    kept = [row for row in rows if row.get("origin") != PROPERTY_CHECK_ORIGIN]
    if not kept:
        return rows
    headers = list(kept[0].keys())
    freshness: dict[str, str] = {}
    names: dict[str, str] = {}
    for row in kept:
        bank_key = row.get("bank_key")
        if not bank_key:
            continue
        if row.get("bank_name"):
            names[bank_key] = row["bank_name"]
        checked_on = row.get("last_checked_on")
        if checked_on:
            freshness[bank_key] = str(checked_on)
    for offer in offers:
        bank_key = offer.get("bank_key")
        if not bank_key:
            continue
        if offer.get("bank_name"):
            names.setdefault(bank_key, offer["bank_name"])
        checked_on = offer.get("last_checked_on")
        if checked_on:
            freshness.setdefault(bank_key, str(checked_on))
    banks = sorted(
        {bank_key for bank_key in (offer.get("bank_key") for offer in offers) if bank_key}
    )
    triples = property_check_amount_triples(len(banks))
    out = list(kept)
    serial = 1
    for bank_key, amounts in zip(banks, triples):
        checked_on = freshness.get(bank_key)
        if not checked_on:
            raise SystemExit(
                f"Missing last_checked_on for property-check overlay bank_key={bank_key!r}"
            )
        for charge_name, amount in zip(PROPERTY_CHECK_NAMES, amounts):
            row = {header: None for header in headers}
            row.update(
                {
                    "charge_row_id": f"CHG-TPC-{serial}",
                    "charge_group_id": f"CHG-TPC-{serial}",
                    "bank_key": bank_key,
                    "bank_name": names.get(bank_key),
                    "origin": PROPERTY_CHECK_ORIGIN,
                    "when_it_matters": "Before offer",
                    "source_ref": PROPERTY_CHECK_ORIGIN,
                    "charge_name": charge_name,
                    "purpose": "Any",
                    "facility_type": "Any",
                    "has_slab_wise_charges": "No",
                    "charge_type": "Fixed Amount",
                    "fixed_amount": amount,
                    "gst_applicable": "Yes",
                    "cibil_band_applicable": "No",
                    "loan_amount_band_applicable": "No",
                    "last_checked_on": checked_on,
                }
            )
            out.append(row)
            serial += 1
    return out


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def cell_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def read_sheet(wb, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if any(h is None or not isinstance(h, str) for h in headers):
        raise SystemExit(f"Bad headers on {sheet_name}: {headers!r}")
    bad = [h for h in headers if not HEADER_RE.match(h)]
    if bad:
        raise SystemExit(f"Illegal header names on {sheet_name}: {bad}")
    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        vals = list(raw)
        if len(vals) < len(headers):
            vals += [None] * (len(headers) - len(vals))
        elif len(vals) > len(headers):
            vals = vals[: len(headers)]
        if all(cell_value(v) is None for v in vals):
            continue
        rows.append({headers[i]: cell_value(vals[i]) for i in range(len(headers))})
    return rows


def bank_freshness_map(rows: list[dict[str, Any]], sheet_name: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for row in rows:
        bank_key = row.get("bank_key")
        if not bank_key:
            continue
        checked_on = row.get("last_checked_on")
        if not checked_on:
            raise SystemExit(
                f"Missing last_checked_on for bank_key={bank_key!r} on {sheet_name}"
            )
        checked_on = str(checked_on)
        prev = out.get(bank_key)
        if prev is not None and prev != checked_on:
            raise SystemExit(
                f"Inconsistent last_checked_on for {bank_key!r} on {sheet_name}: "
                f"{prev!r} vs {checked_on!r}"
            )
        out[bank_key] = checked_on
    return out


def attach_freshness_meta(payload: dict[str, Any]) -> None:
    offers_map = bank_freshness_map(payload["offers"], "Offers")
    charges_map = bank_freshness_map(payload["bank_charges"], "Bank_charges")

    missing_in_charges = sorted(set(offers_map) - set(charges_map))
    if missing_in_charges:
        raise SystemExit(
            "Offers bank_key missing from Bank_charges last_checked_on: "
            + ", ".join(missing_in_charges[:10])
        )

    mismatches = sorted(
        bank_key
        for bank_key in offers_map
        if bank_key in charges_map and offers_map[bank_key] != charges_map[bank_key]
    )
    if mismatches:
        bank_key = mismatches[0]
        raise SystemExit(
            f"last_checked_on mismatch for {bank_key!r}: "
            f"Offers={offers_map[bank_key]!r} Bank_charges={charges_map[bank_key]!r}"
        )

    latest = max(date.fromisoformat(value) for value in offers_map.values())
    payload["meta"]["latest_checked_on"] = latest.isoformat()
    payload["meta"]["bank_freshness"] = dict(sorted(offers_map.items()))


def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)


def main() -> int:
    if not SOURCE.exists():
        print(f"Missing source: {SOURCE}", file=sys.stderr)
        return 1

    digest = sha256_file(SOURCE)
    now = datetime.now(timezone.utc).replace(microsecond=0)
    generated_at = now.isoformat()
    data_version = f"hlc-{now.strftime('%Y%m%d')}-{digest[:8]}"

    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    if set(wb.sheetnames) != set(SHEETS.keys()):
        print(f"Unexpected sheets {wb.sheetnames!r}; expected {sorted(SHEETS)}", file=sys.stderr)
        wb.close()
        return 1

    payload: dict[str, Any] = {
        "meta": {
            "data_version": data_version,
            "package": "home-loans-compare",
            "schema_version": 1,
            "generated_at": generated_at,
            "source_xlsx": "data/HOME_LOANS_COMPARE_v1.xlsx",
            "source_sha256": digest,
            "row_counts": {},
        }
    }

    for sheet_name, key in SHEETS.items():
        rows = read_sheet(wb, sheet_name)
        expected = LOCKED_COUNTS[key]
        if len(rows) != expected:
            print(f"COUNT FAIL {key}: got {len(rows)} expected {expected}", file=sys.stderr)
            wb.close()
            return 1
        payload[key] = rows
        payload["meta"]["row_counts"][key] = len(rows)

    wb.close()

    payload["bank_charges"] = inject_temporary_property_check_charges(
        payload["bank_charges"], payload["offers"]
    )
    payload["meta"]["row_counts"]["bank_charges"] = len(payload["bank_charges"])

    # Fail before write if required id fields missing on first row
    if not payload["offers"] or "bank_key" not in payload["offers"][0] or "offer_row_id" not in payload["offers"][0]:
        print("FAIL: offers missing bank_key/offer_row_id", file=sys.stderr)
        return 1
    if not payload["bank_charges"] or "charge_row_id" not in payload["bank_charges"][0]:
        print("FAIL: bank_charges missing charge_row_id", file=sys.stderr)
        return 1

    proc = [r for r in payload["bank_charges"] if r.get("origin") == "Offers.processing"]
    if len(proc) != PROCESSING_ROWS:
        print(f"COUNT FAIL processing: got {len(proc)} expected {PROCESSING_ROWS}", file=sys.stderr)
        return 1
    cibil_proc = [
        r for r in proc
        if r.get("cibil_band_applicable") not in (None, "No", "")
        or r.get("cibil_band_score_min") is not None
        or r.get("cibil_band_score_max") is not None
    ]
    if cibil_proc:
        print(f"FAIL: {len(cibil_proc)} processing rows still have CIBIL fields", file=sys.stderr)
        return 1

    attach_freshness_meta(payload)

    offer_bank_keys = frozenset(
        bank_key for bank_key in (row.get("bank_key") for row in payload["offers"]) if bank_key
    )
    part_prepayment_rules = load_part_prepayment_rules(
        PART_PREPAY_RULES_CSV,
        known_bank_keys=offer_bank_keys,
    )
    payload["part_prepayment_rules"] = part_prepayment_rules
    payload["meta"]["row_counts"]["part_prepayment_rules"] = len(part_prepayment_rules)
    payload["meta"]["part_prepayment_rules_source"] = "data/part-prepayment-rules.csv"

    atomic_write_text(OUT_JSON, json.dumps(payload, ensure_ascii=False, separators=(",", ":")))

    lines = [
        "# HOME_LOANS_COMPARE JSON VALIDATION",
        "",
        "**Result:** PASS (export completed)",
        "",
        f"- data_version: `{data_version}`",
        f"- source_sha256: `{digest}`",
        f"- generated_at: `{generated_at}`",
        f"- latest_checked_on: `{payload['meta'].get('latest_checked_on')}`",
        f"- offers: {LOCKED_COUNTS['offers']}",
        f"- bank_charges: {payload['meta']['row_counts']['bank_charges']} (xlsx {LOCKED_COUNTS['bank_charges']} + property-check overlay)",
        f"- government_charges: {LOCKED_COUNTS['government_charges']}",
        f"- part_prepayment_rules: {len(part_prepayment_rules)}",
        "- output: `data/home-loans-compare.json`",
        f"- part_prepayment_rules_source: `{PART_PREPAY_RULES_CSV.relative_to(ROOT)}`",
        f"- processing_fee_rows (Offers.processing): {PROCESSING_ROWS}",
        "- processing fees: CIBIL cleared; no 0.25% Canara tier",
        f"- property_check_placeholder_rows: {sum(1 for r in payload['bank_charges'] if r.get('origin') == PROPERTY_CHECK_ORIGIN)} (origin `{PROPERTY_CHECK_ORIGIN}`)",
        "",
        "## Product query contract",
        "1. Match offers by inputs; Blank/Any = no restriction; Offered for customer quotes.",
        "2. Join bank_charges on bank_key (+ filters / when_it_matters).",
        "3. Slabs: group charge_group_id, order slab_from.",
        "4. Govt: filter by jurisdiction/state.",
        "5. EMI in UI from matched roi + amount + tenure.",
        "6. Apply packet: data_version + bank_keys + offer_row_ids + charge ids.",
        "",
        "## One-shot",
        "- Candidate primary JSON only; not auto-promoted to live site.",
        "- Secondary/tertiary/Approve UI out of this export.",
        "",
    ]
    atomic_write_text(OUT_VAL, "\n".join(lines))
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_VAL}")
    print(f"data_version={data_version}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

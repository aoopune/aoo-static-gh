#!/usr/bin/env python3
"""
Convert HOME LOANS - OFFERS (5).xlsx → data/HOME_LOANS_COMPARE_v1.xlsx
Read-only source. Zero drift vs locked plan. Binary pass/fail.
"""
from __future__ import annotations

import hashlib
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Paths / locked hash / counts
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "HOME LOANS - OFFERS (5).xlsx"
OUT_XLSX = ROOT / "data" / "HOME_LOANS_COMPARE_v1.xlsx"
OUT_VAL = ROOT / "data" / "HOME_LOANS_COMPARE_v1_VALIDATION.md"

LOCKED_SHA256 = "ca920fd36e3a6f4053aea46ccd6161a771cfd4d61a6e5733a41d524fd2e7a843"
LOCKED = {
    "filled_before_drop": 1505,
    "empty_skipped": 1000,
    "nainital_dropped": 7,
    "population": 1498,
    "offers": 1149,
    "other_charges": 834,
    "processing": 1183,
    "overdue": 101,
    "slab": 76,
    "prepayment": 87,
    "prepay_keys": 83,
    "government": 18,  # Loan Agreement Stamp Duty rows removed by curator
}

NAINITAL_DROP = {1363, 1379, 1395, 1397, 1399, 1401, 1403}
NAINITAL_KEEP = {1362, 1378, 1394, 1396, 1398, 1400, 1402}

HEADER_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]*$")
ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
TOL = 1e-12

# ---------------------------------------------------------------------------
# Column lists (plan §§3–5)
# ---------------------------------------------------------------------------
OFFERS_RATE_KEY = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type", "occupation",
    "borrower_category", "roi_basis", "cibil_band_applicable", "cibil_band_score_min",
    "cibil_band_score_max", "loan_amount_band_applicable", "loan_amount_min",
    "loan_amount_max", "tenure_band_applicable", "tenure_months_min", "tenure_months_max",
    "benchmark_type", "benchmark_rate", "markup", "business strategic premium", "base_rate",
    "credit_risk_premium", "strategic_premium", "fixed_rate_premium", "discount", "roi",
    "women_benefit_applicable", "women_discount", "women_roi",
    "green_house_benefit_applicable", "if_green_house_benefit_applicable_yes_then_scheme",
    "green_house_discount", "green_roi", "insurance_pricing_applicable",
    "insurance_pricing_rule", "insurance_adjustment", "insurance_roi", "roi_availability",
    "internal_scorerating", "req_amount_min", "req_amount_max",
    "req_repayment_tenure_months_min", "req_repayment_tenure_months_max", "age_min", "age_max",
]

OFFERS_OUT_HEADERS = [
    "bank_name", "bank_type", "scheme", "purpose", "facility_type", "rate_type", "occupation",
    "borrower_category", "roi_basis", "cibil_band_applicable", "cibil_band_score_min",
    "cibil_band_score_max", "cibil_score_status", "loan_amount_band_applicable",
    "loan_amount_min", "loan_amount_max", "tenure_band_applicable", "tenure_months_min",
    "tenure_months_max", "benchmark_type", "benchmark_rate", "markup",
    "business_strategic_premium", "base_rate", "credit_risk_premium", "strategic_premium",
    "fixed_rate_premium", "discount", "roi", "women_benefit_applicable", "women_discount",
    "women_roi", "green_house_benefit_applicable", "green_house_scheme",
    "green_house_discount", "green_roi", "insurance_pricing_applicable",
    "insurance_pricing_rule", "insurance_adjustment", "insurance_roi", "roi_availability",
    "internal_score_rating", "req_amount_min", "req_amount_max",
    "req_repayment_tenure_months_min", "req_repayment_tenure_months_max", "age_min",
    "age_max", "bank_key", "offer_row_id", "source_ref",
]

# DFS / RBI classification (Public = 12 PSBs; rest Private incl. IDBI, J&K, Nainital)
PUBLIC_BANK_KEYS = frozenset({
    "bank of baroda",
    "bank of india",
    "bank of maharashtra",
    "canara bank",
    "central bank of india",
    "indian bank",
    "indian overseas bank",
    "punjab and sind bank",
    "punjab national bank",
    "state bank of india",
    "uco bank",
    "union bank of india",
})

OFFERS_RENAME = {
    "business strategic premium": "business_strategic_premium",
    "if_green_house_benefit_applicable_yes_then_scheme": "green_house_scheme",
    "internal_scorerating": "internal_score_rating",
}

FORBIDDEN_OFFERS_FEE = [
    "processing_fee_calc_type", "processing_fee_percent", "processing_fee_flat_amount",
    "processing_fee_min_amount", "processing_fee_max_amount", "processing_fee_valid_from",
    "processing_fee_valid_till", "overdue_charge_calc_type", "overdue_charge_percent_per_annum",
    "overdue_charge_flat_amount", "overdue_charge_min_amount", "overdue_charge_max_amount",
    "overdue_grace_days", "overdue_whichever_higher", "overdue_days_min", "overdue_days_max",
    "overdue_loan_amount_min", "overdue_loan_amount_max", "overdue_tenure_months_min",
    "overdue_tenure_months_max", "overdue_charge_base", "prepayment_applicable",
    "self_prepayment_percent", "self_prepayment_charge_base", "takeover_prepayment_percent",
    "takeover_prepayment_charge_base",
]

OTHER_57 = [
    "bank_name", "charge_name", "property_valuation_scope", "purpose", "employment_type",
    "facility_type", "has_slab_wise_charges", "slab_from", "slab_to", "slab_basis",
    "charge_type", "fixed_amount", "fixed_amount_at_branch",
    "fixed_amount_at_net_mobile_banking", "fixed_amount_unit", "fixed_amount_per_1000_rs",
    "fixed_amount_per_lakh_or_part", "percentage", "percentage_per_10_lakh_beyond_slab_from",
    "percentage_per_annum", "percentage_base_value", "charge_unit", "charge_min",
    "charge_max", "charge_max_unit", "actuals_in_addition_to_charge", "has_grace_period",
    "grace_period_days", "grace_period_months", "charge_frequency_per_financial_year",
    "charge_frequency_other", "out_of_pocket_expenses_additional", "gst_applicable",
    "refundable_if_not_sanctioned", "facility_conversion_from", "facility_conversion_to",
    "facility_conversion_from_name", "facility_conversion_to_name",
    "interest_rate_repricing_type", "interest_rate_repricing_from",
    "interest_rate_repricing_to", "benchmark_switch_from", "benchmark_switch_to",
    "interest_rate_type_switch_from", "interest_rate_type_switch_to", "customer_type",
    "charged_for_physical_copy", "charged_for_digital_copy",
    "charged_for_original_copy_or_first_issue", "charge_by_area",
    "utilisation_below_per_quarter", "note_1", "note_2", "special_rule", "exemption_1",
    "exemption_2", "exemption_3",
]

OFFERS_FILTER_EXTRA = [
    "scheme", "rate_type", "occupation", "borrower_category", "cibil_band_applicable",
    "cibil_band_score_min", "cibil_band_score_max", "loan_amount_band_applicable",
    "loan_amount_min", "loan_amount_max", "tenure_band_applicable", "tenure_months_min",
    "tenure_months_max",
]

BANK_META = [
    "charge_row_id", "charge_group_id", "bank_key", "origin", "when_it_matters",
    "source_ref", "percentage_unit", "valid_from", "valid_till",
]

BANK_HEADERS = BANK_META + OTHER_57 + OFFERS_FILTER_EXTRA

GOVT_SRC_TO_OUT = [
    ("charge_name", "charge_name"),
    ("jurisdiction_state", "jurisdiction_state"),
    ("jurisdiction_level", "jurisdiction_level"),
    ("mortgage_type", "mortgage_type"),
    ("Calculation_Method", "calculation_method"),
    ("slab_wise_charges", "slab_wise_charges"),
    ("Slab_Min_INR", "slab_min_inr"),
    ("Slab_Max_INR", "slab_max_inr"),
    ("Slab_basis", "slab_basis"),
    ("Flat_Amount_INR", "flat_amount_inr"),
    ("Percentage", "percentage"),
    ("Min_Amount_INR", "min_amount_inr"),
    ("Max_Amount_INR", "max_amount_inr"),
    ("Percentage_Calculated_On", "percentage_calculated_on"),
    ("Is_Consortium_Lending", "is_consortium_lending"),
    ("Is_Top_Up_Or_Further_Charge", "is_top_up_or_further_charge"),
    ("Is_Collateral_Security", "is_collateral_security"),
    ("Modt_Stamp_Already_Paid", "modt_stamp_already_paid"),
    ("Is_Physical_Filing", "is_physical_filing"),
    ("GST_Applicable", "gst_applicable"),
    ("Note", "note"),
]
GOVT_HEADERS = [o for _, o in GOVT_SRC_TO_OUT] + [
    "charge_row_id", "source_ref", "percentage_unit",
]

PROC_KEY = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type", "occupation",
    "borrower_category", "cibil_band_applicable", "cibil_band_score_min",
    "cibil_band_score_max", "loan_amount_band_applicable", "loan_amount_min",
    "loan_amount_max", "tenure_band_applicable", "tenure_months_min",
    "tenure_months_max", "processing_fee_calc_type", "processing_fee_percent",
    "processing_fee_flat_amount", "processing_fee_min_amount", "processing_fee_max_amount",
    "processing_fee_valid_from", "processing_fee_valid_till",
]
PROC_MATCH_NO_AMT = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type", "occupation",
    "borrower_category", "cibil_band_applicable", "cibil_band_score_min",
    "cibil_band_score_max", "loan_amount_band_applicable", "loan_amount_min",
    "loan_amount_max", "tenure_band_applicable", "tenure_months_min",
    "tenure_months_max", "processing_fee_valid_from", "processing_fee_valid_till",
]

OD_KEY = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type",
    "overdue_charge_calc_type", "overdue_charge_percent_per_annum",
    "overdue_charge_flat_amount", "overdue_charge_min_amount", "overdue_charge_max_amount",
    "overdue_grace_days", "overdue_whichever_higher", "overdue_days_min", "overdue_days_max",
    "overdue_loan_amount_min", "overdue_loan_amount_max", "overdue_tenure_months_min",
    "overdue_tenure_months_max", "overdue_charge_base",
]

PRE_KEY = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type",
    "prepayment_applicable", "self_prepayment_percent", "self_prepayment_charge_base",
    "takeover_prepayment_percent", "takeover_prepayment_charge_base",
]

OC_SLAB_GROUP_KEY = [
    "bank_name", "charge_name", "purpose", "facility_type", "employment_type",
    "customer_type", "charge_by_area", "interest_rate_type_switch_from",
    "interest_rate_type_switch_to", "interest_rate_repricing_from",
    "interest_rate_repricing_to", "facility_conversion_from", "facility_conversion_to",
    "benchmark_switch_from", "benchmark_switch_to", "charge_unit", "note_1",
]

FIXED_BLANK_PREMIUM_ROWS = set(range(1052, 1055)) | set(range(1131, 1140))

# ---------------------------------------------------------------------------
# when_it_matters (plan §11)
# ---------------------------------------------------------------------------
BEFORE_KW = [
    k.strip().lower()
    for k in """valuation, legal, title search, title search report, documentation,
cibil, cic fees, cic), credit information, credit opinion,
encumbrance, non-encumbrance, inspection, field investigation,
technical inspection, administrative charges, service charges,
advocate, pre-credit""".replace("\n", " ").split(",")
    if k.strip()
]
AFTER_KW = [
    k.strip().lower()
    for k in """switch, repric, benchmark, conversion, bounce, return, dishonour,
overdue, penal, npa, overdrawn, non util, non-util, commitment,
failed, nach, ecs, standing instruction, no objection, no due, no dues,
statement of account, amortis, certificate, document copy, document retrieval,
passbook, solvency, list of documents, modification, amendment, revalidation,
deviation, cancellation, rebooking, reschedul, rephas, property swap,
partial property, renewal fee, mandate, swap, neft, rtgs, imps, collection,
cheque book, folio, stop payment, prepayment, foreclos, max saver""".replace(
        "\n", " "
    ).split(",")
    if k.strip()
]

BOTH_OVERRIDE = {
    "Non-Encumbrance Certificate Charges": "Before offer",
    "Penal Charge - Non-execution of Documentation": "After offer",
}

EXPLICIT_BEFORE = set(
    """Confidential Opinion Report Charge
External Due Diligence Charge
Equitable Mortgage Creation Charge
Property Valuation Report Charges
Property Valuation Report Charges - Agricultural land
Property Valuation Report Charges - Local
Property Valuation Report Charges - Outstation
Documentation Charges
Documentation and Inspection Charge (Composite)
Loan Documentation Charges
Digital Documentation Charge
Credit Information Report (CIC) Charges
Credit Information Report (CIC) Charges - Copy
CIC Fees
CIBIL Report Charge
CIBIL Charges
CIBIL Report Retrieval
CIBIL Report Pulling Charge
CIBIL/CRIF Consumer Report Charge
CIBIL/CRIF Corporate Report Charge
Inspection Charges
Property Inspection Charge
Property Inspection Fee
Property Inspection Fee (Beyond Local Limits)
Technical Inspection Charges — Local
Technical Inspection Charges — Outstation
Field Investigation Charge
Security Inspection / Verification Charge
Legal Opinion Fees
Legal Charges
Legal Audit Fee
Legal and Valuation Charges
Legal & Technical Charges
Legal / Technical Verification Fee
Legal and Pre-Credit Inspection / Verification Charge
Legal Fee to Advocate
Title Search Report Fees
Non-Encumbrance Certificate Charges
Administrative Charges
Service charges""".splitlines()
)

EXPLICIT_AFTER = set(
    """API Integration Charge
Account Handling Charge
Additional Housing Loan Service Charge at Disbursal
Administrative Fee for Non-Auto Debit cases
Breach of Construction Timeline Charges
Cash Deposit Amount Charges
Cash Deposit Frequency Charges
Cash EMI Payment Charges
Ceding of Charge on Security including Pari Passu Charge
Cheque Leaf Above Free Limit
Construction Delay Charge
Delay in Creation of EM Charge Charges
Delay in Delay in renewal/review or Non-Renewal / Review of Facility
Drawdown Failure Charge
Duplicate Loan Agreement Copy
EMI Cycle Change Charge
EMI Cycle Date Change Charge
Escrow Account Charge
Exchange / Remittance Charges on Term Loan Disbursement Drafts
Hard Copy of Other Documents / Letters
Incidental Charges
Incidental Loan Closure Charge
Interest on amount utilized above Operating Limit (DOD)
Letter of Acknowledgement of Debt (LAD)
Loose Cheque Leaf Issuance
Max Saver Charge Rule
NeSL Digital Document Storage Charge
Non-Adherence to Material Terms and Conditions
Non-Compliance of Terms of Sanction (Plot + Construction)
Non-Maintenance of Mode of Payment (NMMP) Charges
Non-Submission of Post Disbursement Documents Charges
Old Records Enquiry Charge
Omni Pay Issuance Charge
Property Document Retention charges
Re-Appraisal Of Loan After 6 Months From Sanction
Re-issuance of PO/DD Charge
Release of Personal Guarantee or Collateral Security
Repayment Mode Change Charges
Repayment Schedule (Physical Copy)
Repayment Schedule (Soft Copy)
Signature Verification / Loan Certification
Signature Verification / Photo Attestation
Stage-wise Completion Report Fee
Substitution of Collateral Security or Personal Guarantee Charges
Term Loan Disbursement Instrument Charge
Term Loan Review Charges
Safe Custody Charges for Title Deeds
Other Services Charges
Other miscellaneous approvals not specified""".splitlines()
)
# Curator locked: former NEEDS_REVIEW names + Safe Custody → After offer.
# Loan Agreement Stamp Duty omitted from Government output (curator deleted).

ALLOWED_NEEDS_REVIEW: set[str] = set()  # none allowed after curator pass

GOVT_SKIP_CHARGE_NAMES = {"Loan Agreement Stamp Duty"}


class ConversionError(Exception):
    pass


class AssertCollector:
    def __init__(self) -> None:
        self.failures: list[str] = []
        self.notes: list[str] = []

    def check(self, name: str, ok: bool, detail: str = "") -> None:
        if not ok:
            self.failures.append(f"FAIL {name}: {detail}")

    def note(self, msg: str) -> None:
        self.notes.append(msg)

    def ok(self) -> bool:
        return not self.failures


def blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def bank_key(name: Any) -> str:
    s = "" if name is None else str(name)
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("&", "and")
    return s


def bank_type_for(name: Any) -> str:
    """Return 'Public' or 'Private' from DFS PSB list; raise if bank_key unknown."""
    key = bank_key(name)
    if not key:
        raise ValueError("blank bank_name for bank_type")
    return "Public" if key in PUBLIC_BANK_KEYS else "Private"


def iso_date(v: Any) -> Any:
    if blank(v):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if ISO_DATE_RE.match(s):
            return s
        # try common parse
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s[:10], fmt).date().isoformat()
            except ValueError:
                continue
        return s
    return str(v)


def num0(v: Any) -> float:
    if blank(v):
        return 0.0
    return float(v)


def row_dict(headers: list[str], values: tuple | list) -> dict[str, Any]:
    vals = list(values)
    if len(vals) < len(headers):
        vals += [None] * (len(headers) - len(vals))
    return {h: vals[i] for i, h in enumerate(headers)}


def key_tuple(d: dict[str, Any], fields: list[str]) -> tuple:
    return tuple(d.get(f) for f in fields)


def empty_bank_row() -> dict[str, Any]:
    return {h: None for h in BANK_HEADERS}


def append_note(row: dict[str, Any], text: str) -> None:
    if blank(row.get("note_2")):
        row["note_2"] = text
    elif blank(row.get("note_1")):
        row["note_1"] = text
    else:
        row["note_2"] = f"{row['note_2']}; {text}"


def classify_when(origin: str, charge_name: Any) -> str:
    if origin == "Offers.processing":
        return "Before offer"
    if origin in ("Offers.overdue", "Offers.prepayment", "Slab_Table"):
        return "After offer"
    name = charge_name if isinstance(charge_name, str) else ("" if charge_name is None else str(charge_name))
    if name in BOTH_OVERRIDE:
        return BOTH_OVERRIDE[name]
    if name in EXPLICIT_BEFORE:
        return "Before offer"
    if name in EXPLICIT_AFTER:
        return "After offer"
    low = name.lower()
    hit_b = any(k in low for k in BEFORE_KW)
    hit_a = any(k in low for k in AFTER_KW)
    if hit_b and hit_a:
        return BOTH_OVERRIDE.get(name, "NEEDS_REVIEW")
    if hit_b and not hit_a:
        return "Before offer"
    if hit_a and not hit_b:
        return "After offer"
    return "NEEDS_REVIEW"


def cibil_status(d: dict[str, Any]) -> str:
    if d.get("cibil_band_applicable") == "No":
        return "Not_Used"
    mn, mx = d.get("cibil_band_score_min"), d.get("cibil_band_score_max")
    if mn == -1 and mx == 0:
        return "No_Score"
    if mn is not None and mx is not None:
        try:
            mnf, mxf = float(mn), float(mx)
            if 0 <= mnf and mxf <= 200:
                return "Thin_File"
        except (TypeError, ValueError):
            pass
    return "Scored"


def normalize_band_flag(flag: Any, mn: Any, mx: Any) -> str:
    if flag in ("Yes", "No"):
        return flag
    if blank(mn) and blank(mx):
        return "No"
    return "Yes"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])


def main() -> int:
    asserts = AssertCollector()
    report: dict[str, Any] = {"sections": []}

    if not SOURCE.exists():
        print(f"Source missing: {SOURCE}", file=sys.stderr)
        return 1

    src_stat = SOURCE.stat()
    digest = sha256_file(SOURCE)
    asserts.check("source_sha256", digest == LOCKED_SHA256, f"got {digest}")
    if digest != LOCKED_SHA256:
        _write_validation(asserts, report, digest, src_stat, None, None, None, {}, abort=True)
        return 1

    wb_src = load_workbook(SOURCE, read_only=True, data_only=True)

    # ----- Offers read -----
    ws_o = wb_src["Offers"]
    offers_headers = [c.value for c in next(ws_o.iter_rows(min_row=1, max_row=1))]
    filled_rows: list[tuple[int, dict[str, Any]]] = []
    empty_count = 0
    for rnum, raw in enumerate(ws_o.iter_rows(min_row=2, values_only=True), start=2):
        d = row_dict(offers_headers, raw)
        if all(blank(d.get(h)) for h in offers_headers):
            empty_count += 1
            continue
        filled_rows.append((rnum, d))

    asserts.check("filled_before_drop", len(filled_rows) == LOCKED["filled_before_drop"], str(len(filled_rows)))
    asserts.check("empty_skipped", empty_count == LOCKED["empty_skipped"], str(empty_count))

    dropped = [r for r, _ in filled_rows if r in NAINITAL_DROP]
    asserts.check("nainital_dropped_set", set(dropped) == NAINITAL_DROP, str(dropped))
    pop = [(r, d) for r, d in filled_rows if r not in NAINITAL_DROP]
    asserts.check("population", len(pop) == LOCKED["population"], str(len(pop)))

    # ----- Extract processing (Block B) -----
    proc_groups: dict[tuple, list[int]] = defaultdict(list)
    for r, d in pop:
        proc_groups[key_tuple(d, PROC_KEY)].append(r)
    asserts.check("processing_keys", len(proc_groups) == LOCKED["processing"], str(len(proc_groups)))

    # ambiguous: same match+dates, different amounts
    ambig_match: dict[tuple, list[tuple]] = defaultdict(list)
    for k in proc_groups:
        ambig_match[tuple(k[i] for i, f in enumerate(PROC_KEY) if f in PROC_MATCH_NO_AMT)].append(k)
    processing_ambiguous: list[dict[str, Any]] = []
    ambig_keys = set()
    for mk, keys in ambig_match.items():
        if len(keys) > 1:
            for k in keys:
                ambig_keys.add(k)
            processing_ambiguous.append({"match": mk, "keys": len(keys), "rows": [proc_groups[k] for k in keys]})

    bank_rows: list[dict[str, Any]] = []
    origin_counters: Counter[str] = Counter()
    seq: dict[str, int] = defaultdict(int)

    def next_id(code: str) -> str:
        seq[code] += 1
        return f"CHG-{code}-{seq[code]}"

    pop_by_row = {r: d for r, d in pop}

    for k, rnums in sorted(proc_groups.items(), key=lambda x: min(x[1])):
        src_row = min(rnums)
        d = pop_by_row[src_row]
        row = empty_bank_row()
        row["charge_row_id"] = next_id("PROC")
        row["charge_group_id"] = row["charge_row_id"]
        row["bank_name"] = d["bank_name"]
        row["bank_key"] = bank_key(d["bank_name"])
        row["origin"] = "Offers.processing"
        row["when_it_matters"] = "Before offer"
        row["source_ref"] = f"Offers!{src_row}"
        row["percentage_unit"] = "fraction"
        row["charge_name"] = "Processing fee"
        row["has_slab_wise_charges"] = "No"
        for f in OFFERS_FILTER_EXTRA + ["purpose", "facility_type"]:
            row[f] = d.get(f)
        # also copy occupation/borrower already in OFFERS_FILTER_EXTRA
        calc = d.get("processing_fee_calc_type")
        if calc == "Percentage":
            row["charge_type"] = "Percentage"
            row["percentage"] = d.get("processing_fee_percent")
            row["percentage_base_value"] = "Sanctioned loan amount"
        elif calc == "Flat":
            row["charge_type"] = "Fixed Amount"
            row["fixed_amount"] = d.get("processing_fee_flat_amount")
        else:
            asserts.check("processing_calc_type", False, f"row {src_row} calc={calc}")
        row["charge_min"] = d.get("processing_fee_min_amount")
        row["charge_max"] = d.get("processing_fee_max_amount")
        row["valid_from"] = iso_date(d.get("processing_fee_valid_from"))
        row["valid_till"] = iso_date(d.get("processing_fee_valid_till"))
        if k in ambig_keys:
            append_note(row, "SOURCE_AMBIGUOUS_PROCESSING_SAME_MATCH")
        bank_rows.append(row)
        origin_counters["Offers.processing"] += 1

    # ----- Extract prepay (Block E) -----
    pre_groups: dict[tuple, list[int]] = defaultdict(list)
    for r, d in pop:
        pre_groups[key_tuple(d, PRE_KEY)].append(r)
    asserts.check("prepay_keys", len(pre_groups) == LOCKED["prepay_keys"], str(len(pre_groups)))

    prepay_contrib: list[dict[str, Any]] = []
    for k, rnums in sorted(pre_groups.items(), key=lambda x: min(x[1])):
        src_row = min(rnums)
        d = pop_by_row[src_row]
        prepay_contrib.append({"key_row": src_row, "contributors": sorted(rnums)})
        filters = {f: d.get(f) for f in ["bank_name", "scheme", "purpose", "facility_type", "rate_type"]}

        def base_prepay_row() -> dict[str, Any]:
            row = empty_bank_row()
            row["bank_name"] = d["bank_name"]
            row["bank_key"] = bank_key(d["bank_name"])
            row["origin"] = "Offers.prepayment"
            row["when_it_matters"] = "After offer"
            row["source_ref"] = f"Offers!{src_row}"
            row["percentage_unit"] = "fraction"
            row["has_slab_wise_charges"] = "No"
            for f, v in filters.items():
                row[f] = v
            return row

        if d.get("prepayment_applicable") == "No":
            row = base_prepay_row()
            row["charge_row_id"] = next_id("PRE")
            row["charge_group_id"] = row["charge_row_id"]
            row["charge_name"] = "Prepayment charges"
            row["charge_type"] = "Fixed Amount"
            row["fixed_amount"] = 0
            row["note_1"] = "Prepayment not charged (prepayment_applicable=No)"
            bank_rows.append(row)
            origin_counters["Offers.prepayment"] += 1
        else:
            self_p, self_b = d.get("self_prepayment_percent"), d.get("self_prepayment_charge_base")
            take_p, take_b = d.get("takeover_prepayment_percent"), d.get("takeover_prepayment_charge_base")
            if not blank(self_p) or not blank(self_b):
                row = base_prepay_row()
                row["charge_row_id"] = next_id("PRE")
                row["charge_group_id"] = row["charge_row_id"]
                row["charge_name"] = "Prepayment charges"
                row["charge_type"] = "Percentage"
                row["percentage"] = self_p
                row["percentage_base_value"] = self_b
                bank_rows.append(row)
                origin_counters["Offers.prepayment"] += 1
            if not blank(take_p) or not blank(take_b):
                row = base_prepay_row()
                row["charge_row_id"] = next_id("PRE")
                row["charge_group_id"] = row["charge_row_id"]
                row["charge_name"] = "Prepayment charges (takeover)"
                row["charge_type"] = "Percentage"
                row["percentage"] = take_p
                row["percentage_base_value"] = take_b
                bank_rows.append(row)
                origin_counters["Offers.prepayment"] += 1

    asserts.check(
        "prepayment_rows",
        origin_counters["Offers.prepayment"] == LOCKED["prepayment"],
        str(origin_counters["Offers.prepayment"]),
    )

    # ----- Extract overdue (Block C) -----
    od_groups: dict[tuple, list[int]] = defaultdict(list)
    slab_deferred_rows: list[int] = []
    for r, d in pop:
        if d.get("overdue_charge_calc_type") == "Slab_Table":
            slab_deferred_rows.append(r)
            continue
        od_groups[key_tuple(d, OD_KEY)].append(r)
    asserts.check("overdue_keys", len(od_groups) == LOCKED["overdue"], str(len(od_groups)))
    asserts.check("slab_deferred_offers_rows", len(slab_deferred_rows) == 6, str(len(slab_deferred_rows)))

    overdue_contrib: list[dict[str, Any]] = []
    for k, rnums in sorted(od_groups.items(), key=lambda x: min(x[1])):
        src_row = min(rnums)
        d = pop_by_row[src_row]
        overdue_contrib.append({"key_row": src_row, "contributors": sorted(rnums)})
        row = empty_bank_row()
        row["charge_row_id"] = next_id("OD")
        row["charge_group_id"] = row["charge_row_id"]
        row["bank_name"] = d["bank_name"]
        row["bank_key"] = bank_key(d["bank_name"])
        row["origin"] = "Offers.overdue"
        row["when_it_matters"] = "After offer"
        row["source_ref"] = f"Offers!{src_row}"
        row["percentage_unit"] = "fraction"
        row["charge_name"] = "Overdue charges"
        for f in ["scheme", "purpose", "facility_type", "rate_type"]:
            row[f] = d.get(f)
        calc = d.get("overdue_charge_calc_type")
        notes: list[str] = []
        if calc == "Percentage":
            row["charge_type"] = "Percentage"
            row["percentage"] = d.get("overdue_charge_percent_per_annum")
            row["percentage_per_annum"] = "Yes"
        elif calc == "As_Per_ROI":
            row["charge_type"] = "Percentage"
            row["percentage"] = d.get("overdue_charge_percent_per_annum")
            row["percentage_per_annum"] = "Yes"
            row["special_rule"] = "As_Per_ROI"
        else:
            asserts.check("overdue_calc", False, f"row {src_row} calc={calc}")
        if not blank(d.get("overdue_charge_flat_amount")):
            row["fixed_amount"] = d.get("overdue_charge_flat_amount")
        row["charge_min"] = d.get("overdue_charge_min_amount")
        row["charge_max"] = d.get("overdue_charge_max_amount")
        if not blank(d.get("overdue_grace_days")):
            row["grace_period_days"] = d.get("overdue_grace_days")
            row["has_grace_period"] = "Yes"
        if not blank(d.get("overdue_loan_amount_min")) or not blank(d.get("overdue_loan_amount_max")):
            row["slab_from"] = d.get("overdue_loan_amount_min")
            row["slab_to"] = d.get("overdue_loan_amount_max")
            row["has_slab_wise_charges"] = "Yes"
            row["slab_basis"] = "Default_Amount"
        else:
            row["has_slab_wise_charges"] = "No"
        row["percentage_base_value"] = d.get("overdue_charge_base")
        if d.get("overdue_whichever_higher") == "Yes":
            notes.append("overdue_whichever_higher=Yes")
        for label, field in [
            ("overdue_days_min", "overdue_days_min"),
            ("overdue_days_max", "overdue_days_max"),
            ("overdue_tenure_months_min", "overdue_tenure_months_min"),
            ("overdue_tenure_months_max", "overdue_tenure_months_max"),
        ]:
            if not blank(d.get(field)):
                notes.append(f"{label}={d.get(field)}")
        if notes:
            row["note_1"] = "; ".join(notes)
        bank_rows.append(row)
        origin_counters["Offers.overdue"] += 1

    # ----- Collapse Offers -----
    rate_groups: dict[tuple, list[int]] = defaultdict(list)
    for r, d in pop:
        rate_groups[key_tuple(d, OFFERS_RATE_KEY)].append(r)
    asserts.check("offers_groups", len(rate_groups) == LOCKED["offers"], str(len(rate_groups)))

    collapse_appendix: list[str] = []
    offers_out: list[dict[str, Any]] = []
    seen_keys: set[tuple] = set()

    for k, rnums in sorted(rate_groups.items(), key=lambda x: min(x[1])):
        asserts.check("unique_collapse_key", k not in seen_keys, str(k)[:80])
        seen_keys.add(k)
        canon = min(rnums)
        d = pop_by_row[canon]
        if len(rnums) > 1:
            collapse_appendix.append(f"OFF-{canon} collapsed from {sorted(rnums)}")

        out: dict[str, Any] = {}
        for h in OFFERS_OUT_HEADERS:
            if h in ("bank_key", "offer_row_id", "source_ref", "cibil_score_status", "bank_type"):
                continue
            src_h = h
            for old, new in OFFERS_RENAME.items():
                if new == h:
                    src_h = old
                    break
            out[h] = d.get(src_h)

        out["loan_amount_band_applicable"] = normalize_band_flag(
            d.get("loan_amount_band_applicable"), d.get("loan_amount_min"), d.get("loan_amount_max")
        )
        out["tenure_band_applicable"] = normalize_band_flag(
            d.get("tenure_band_applicable"), d.get("tenure_months_min"), d.get("tenure_months_max")
        )
        out["cibil_score_status"] = cibil_status(d)
        out["bank_key"] = bank_key(d.get("bank_name"))
        out["bank_type"] = bank_type_for(d.get("bank_name"))
        out["offer_row_id"] = f"OFF-{canon}"
        out["source_ref"] = f"Offers!{canon}"

        # ROI QA
        try:
            base_expected = num0(d.get("benchmark_rate")) + num0(d.get("markup")) + num0(
                d.get("business strategic premium")
            )
            asserts.check(
                f"roi_qa_base_{canon}",
                abs(num0(d.get("base_rate")) - base_expected) <= TOL,
                f"base={d.get('base_rate')} expected={base_expected}",
            )
            roi_expected = (
                num0(d.get("base_rate"))
                + num0(d.get("credit_risk_premium"))
                + num0(d.get("strategic_premium"))
                + num0(d.get("fixed_rate_premium"))
                + num0(d.get("discount"))
            )
            asserts.check(
                f"roi_qa_roi_{canon}",
                abs(num0(d.get("roi")) - roi_expected) <= TOL,
                f"roi={d.get('roi')} expected={roi_expected}",
            )
            if d.get("women_benefit_applicable") == "Yes":
                asserts.check(
                    f"roi_qa_women_{canon}",
                    abs(num0(d.get("women_roi")) - (num0(d.get("roi")) - num0(d.get("women_discount")))) <= TOL,
                    "",
                )
            if d.get("green_house_benefit_applicable") == "Yes":
                asserts.check(
                    f"roi_qa_green_{canon}",
                    abs(num0(d.get("green_roi")) - (num0(d.get("roi")) - num0(d.get("green_house_discount"))))
                    <= TOL,
                    "",
                )
            if d.get("insurance_pricing_applicable") == "Yes":
                asserts.check(
                    f"roi_qa_ins_{canon}",
                    abs(num0(d.get("insurance_roi")) - (num0(d.get("roi")) + num0(d.get("insurance_adjustment"))))
                    <= TOL,
                    "",
                )
            if d.get("rate_type") == "Floating":
                asserts.check(
                    f"roi_qa_float_prem_{canon}",
                    blank(d.get("fixed_rate_premium")),
                    f"fixed_rate_premium={d.get('fixed_rate_premium')}",
                )
        except (TypeError, ValueError) as e:
            asserts.check(f"roi_qa_numeric_{canon}", False, str(e))

        offers_out.append(out)

    asserts.check("offers_count", len(offers_out) == LOCKED["offers"], str(len(offers_out)))
    asserts.check(
        "offers_flags",
        all(o["loan_amount_band_applicable"] in ("Yes", "No") and o["tenure_band_applicable"] in ("Yes", "No") for o in offers_out),
        "",
    )
    asserts.check(
        "cibil_status_all",
        all(o["cibil_score_status"] in ("Not_Used", "No_Score", "Thin_File", "Scored") for o in offers_out),
        "",
    )
    asserts.check("bank_key_all_offers", all(not blank(o["bank_key"]) for o in offers_out), "")
    asserts.check(
        "bank_type_all_offers",
        all(o["bank_type"] in ("Public", "Private") for o in offers_out),
        "",
    )
    offers_bank_keys = {o["bank_key"] for o in offers_out}
    asserts.check(
        "bank_type_public_count",
        len(offers_bank_keys & PUBLIC_BANK_KEYS) == 12,
        str(sorted(offers_bank_keys & PUBLIC_BANK_KEYS)),
    )
    asserts.check(
        "bank_type_private_count",
        len(offers_bank_keys - PUBLIC_BANK_KEYS) == 21,
        str(sorted(offers_bank_keys - PUBLIC_BANK_KEYS)),
    )
    asserts.check(
        "no_fee_cols",
        not any(f in OFFERS_OUT_HEADERS for f in FORBIDDEN_OFFERS_FEE),
        "",
    )
    for old, new in OFFERS_RENAME.items():
        asserts.check(f"rename_{new}", new in OFFERS_OUT_HEADERS and old not in OFFERS_OUT_HEADERS, "")

    # ----- Government -----
    ws_g = wb_src["Govt. Charges"]
    govt_headers = [c.value for c in next(ws_g.iter_rows(min_row=1, max_row=1))]
    govt_out: list[dict[str, Any]] = []
    govt_pct_log: list[str] = []
    gseq = 0
    for rnum, raw in enumerate(ws_g.iter_rows(min_row=2, values_only=True), start=2):
        d = row_dict(govt_headers, raw)
        if all(blank(v) for v in d.values()):
            continue
        if d.get("charge_name") in GOVT_SKIP_CHARGE_NAMES:
            continue
        gseq += 1
        out = {h: None for h in GOVT_HEADERS}
        for src_h, out_h in GOVT_SRC_TO_OUT:
            val = d.get(src_h)
            if out_h == "percentage" and not blank(val):
                govt_pct_log.append(f"Govt!{rnum} {d.get('charge_name')}: {val} → {float(val)/100}")
                val = float(val) / 100.0
            out[out_h] = val
        out["charge_row_id"] = f"CHG-GOV-{gseq}"
        out["source_ref"] = f"Govt. Charges!{rnum}"
        out["percentage_unit"] = "fraction"
        govt_out.append(out)
    asserts.check("govt_count", len(govt_out) == LOCKED["government"], str(len(govt_out)))
    asserts.check(
        "govt_no_loan_agreement_stamp",
        not any(r.get("charge_name") == "Loan Agreement Stamp Duty" for r in govt_out),
        "",
    )

    # ----- Other charges Block A -----
    ws_oc = wb_src["Other charges"]
    oc_headers = [c.value for c in next(ws_oc.iter_rows(min_row=1, max_row=1))]
    asserts.check("other_57_headers", oc_headers == OTHER_57, f"got {len(oc_headers)}")

    oc_rows_src: list[tuple[int, dict[str, Any]]] = []
    for rnum, raw in enumerate(ws_oc.iter_rows(min_row=2, values_only=True), start=2):
        d = row_dict(oc_headers, raw)
        if all(blank(v) for v in d.values()):
            continue
        oc_rows_src.append((rnum, d))
    asserts.check("other_count", len(oc_rows_src) == LOCKED["other_charges"], str(len(oc_rows_src)))

    # slab groups first pass
    slab_group_map: dict[tuple, str] = {}
    oc_bank: list[dict[str, Any]] = []
    for rnum, d in oc_rows_src:
        row = empty_bank_row()
        cid = next_id("OC")
        row["charge_row_id"] = cid
        for h in OTHER_57:
            row[h] = d.get(h)
        # purpose alias
        if row.get("purpose") == "Home Loan":
            row["purpose"] = "Regular Home Loan"
            append_note(row, "purpose_normalized_from=Home Loan")
        row["bank_key"] = bank_key(row.get("bank_name"))
        row["origin"] = "Other charges"
        row["source_ref"] = f"Other charges!{rnum}"
        row["percentage_unit"] = "fraction"
        row["when_it_matters"] = classify_when("Other charges", row.get("charge_name"))
        if row.get("has_slab_wise_charges") == "Yes":
            gk = key_tuple(d, OC_SLAB_GROUP_KEY)
            if gk not in slab_group_map:
                slab_group_map[gk] = cid
            row["charge_group_id"] = slab_group_map[gk]
        else:
            row["charge_group_id"] = cid
        oc_bank.append(row)
        origin_counters["Other charges"] += 1

    # prepend Other charges before Offers-derived? Plan order: A then B C D E
    # We already built B/C/E into bank_rows; rebuild order: A + B + C + D + E
    offers_derived = bank_rows
    bank_rows = oc_bank + [r for r in offers_derived if r["origin"] == "Offers.processing"]
    bank_rows += [r for r in offers_derived if r["origin"] == "Offers.overdue"]

    # ----- Slab_Table Block D -----
    ws_s = wb_src["Slab_Table"]
    slab_headers = [c.value for c in next(ws_s.iter_rows(min_row=1, max_row=1))]
    slab_out: list[dict[str, Any]] = []
    slab_group_id = None
    for rnum, raw in enumerate(ws_s.iter_rows(min_row=2, values_only=True), start=2):
        d = row_dict(slab_headers, raw)
        if all(blank(v) for v in d.values()):
            continue
        row = empty_bank_row()
        row["charge_row_id"] = next_id("SLAB")
        if slab_group_id is None:
            slab_group_id = row["charge_row_id"]
        row["charge_group_id"] = slab_group_id
        row["bank_name"] = d.get("Bank")
        row["bank_key"] = bank_key(d.get("Bank"))
        row["origin"] = "Slab_Table"
        row["when_it_matters"] = "After offer"
        row["source_ref"] = f"Slab_Table!{rnum}"
        row["percentage_unit"] = "fraction"
        row["charge_name"] = "Overdue charges"
        row["has_slab_wise_charges"] = "Yes"
        row["slab_basis"] = "Default_Amount"
        row["slab_from"] = d.get("Default_Amount_Min")
        row["slab_to"] = d.get("Default_Amount_Max")
        row["fixed_amount"] = d.get("Overdue_Flat_Amount")
        row["charge_type"] = "Fixed Amount"
        row["charge_unit"] = "Instance"
        row["special_rule"] = "Slab_Table"
        row["purpose"] = "Any"
        row["facility_type"] = "Any"
        slab_out.append(row)
        origin_counters["Slab_Table"] += 1
    asserts.check("slab_count", len(slab_out) == LOCKED["slab"], str(len(slab_out)))
    bank_rows += slab_out
    bank_rows += [r for r in offers_derived if r["origin"] == "Offers.prepayment"]

    # Re-apply when_it_matters for all (origin already set for moved rows)
    for row in bank_rows:
        row["when_it_matters"] = classify_when(row["origin"], row.get("charge_name"))

    # Sort slabs within each charge_group_id
    by_group: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for i, row in enumerate(bank_rows):
        by_group[row["charge_group_id"]].append(row)

    def slab_sort_key(r: dict[str, Any]):
        sf = r.get("slab_from")
        if blank(sf):
            return (1, 0)
        try:
            return (0, float(sf))
        except (TypeError, ValueError):
            return (0, 0)

    # rebuild preserving non-slab order but sorting within groups that have slabs
    group_sorted_ids = set()
    for gid, rows in by_group.items():
        if any(r.get("has_slab_wise_charges") == "Yes" for r in rows) and len(rows) > 1:
            rows_sorted = sorted(rows, key=slab_sort_key)
            # check non-decreasing
            prev = None
            for r in rows_sorted:
                sf = r.get("slab_from")
                if blank(sf):
                    continue
                try:
                    cur = float(sf)
                except (TypeError, ValueError):
                    continue
                if prev is not None:
                    asserts.check(f"slab_order_{gid}", cur >= prev, f"{prev} -> {cur}")
                prev = cur
            group_sorted_ids.add(gid)
            by_group[gid] = rows_sorted

    # rebuild bank_rows: walk original order of groups by first appearance
    seen_g = set()
    new_bank: list[dict[str, Any]] = []
    for row in bank_rows:
        gid = row["charge_group_id"]
        if gid in seen_g:
            continue
        seen_g.add(gid)
        if gid in group_sorted_ids:
            new_bank.extend(by_group[gid])
        else:
            new_bank.extend(by_group[gid])
    bank_rows = new_bank

    # Origin counts
    for origin, exp_key in [
        ("Other charges", "other_charges"),
        ("Offers.processing", "processing"),
        ("Offers.overdue", "overdue"),
        ("Slab_Table", "slab"),
        ("Offers.prepayment", "prepayment"),
    ]:
        got = sum(1 for r in bank_rows if r["origin"] == origin)
        asserts.check(f"count_{origin}", got == LOCKED[exp_key], f"got {got}")

    # when_it_matters assert
    nr_names = sorted(
        {
            r.get("charge_name")
            for r in bank_rows
            if r.get("when_it_matters") == "NEEDS_REVIEW"
        }
    )
    asserts.check(
        "needs_review_names",
        len(nr_names) == 0 and set(nr_names) <= ALLOWED_NEEDS_REVIEW,
        f"got {nr_names}",
    )
    asserts.check(
        "when_values",
        all(r.get("when_it_matters") in ("Before offer", "After offer", "NEEDS_REVIEW") for r in bank_rows),
        "",
    )

    # Naming headers
    for sheet_name, headers in [
        ("Offers", OFFERS_OUT_HEADERS),
        ("Bank_charges", BANK_HEADERS),
        ("Government_charges", GOVT_HEADERS),
    ]:
        for h in headers:
            asserts.check(f"naming_{sheet_name}_{h}", bool(HEADER_RE.match(h)), h)

    # Forbidden parallel fee columns
    forbidden_bank = [h for h in BANK_HEADERS if h.startswith("processing_fee_") or "prepayment_" in h or h.startswith("overdue_charge_")]
    asserts.check("no_parallel_fee_cols", not forbidden_bank, str(forbidden_bank))

    # purpose Home Loan gone
    asserts.check(
        "no_purpose_home_loan",
        not any(r.get("purpose") == "Home Loan" for r in bank_rows),
        "",
    )

    # dates ISO
    for r in bank_rows:
        for f in ("valid_from", "valid_till"):
            v = r.get(f)
            if not blank(v):
                asserts.check(f"iso_{r.get('charge_row_id')}_{f}", bool(ISO_DATE_RE.match(str(v))), str(v))

    # Spot checks
    def find_oc(ref: str) -> dict[str, Any] | None:
        for r in bank_rows:
            if r.get("source_ref") == ref:
                return r
        return None

    # Axis switch 13-14
    a13, a14 = find_oc("Other charges!13"), find_oc("Other charges!14")
    asserts.check("spot_axis_13", a13 is not None and a13.get("percentage") == 0.01 and a13.get("charge_min") == 10000, str(a13))
    asserts.check("spot_axis_14", a14 is not None and abs(float(a14.get("percentage") or 0) - 0.02) < 1e-12, str(a14 and a14.get("percentage")))

    # Axis repricing 16-19
    slabs = [find_oc(f"Other charges!{n}") for n in range(16, 20)]
    expected_fixed = [1000, 2000, 3000, 5000]
    asserts.check(
        "spot_axis_repricing",
        all(s and s.get("fixed_amount") == exp for s, exp in zip(slabs, expected_fixed)),
        str([s.get("fixed_amount") if s else None for s in slabs]),
    )

    # BoB ECS 51-56
    for n in range(51, 57):
        asserts.check(f"spot_bob_{n}", find_oc(f"Other charges!{n}") is not None, "")

    # BoM 75-78
    for n in range(75, 79):
        asserts.check(f"spot_bom_{n}", find_oc(f"Other charges!{n}") is not None, "")

    # BoI CIC 62-63
    b62, b63 = find_oc("Other charges!62"), find_oc("Other charges!63")
    asserts.check("spot_boi_62", b62 is not None and b62.get("fixed_amount") == 50, str(b62))
    asserts.check("spot_boi_63", b63 is not None and b63.get("fixed_amount") == 200, str(b63))

    # PNB processing from Offers 2 vs 3
    pnb_proc = [r for r in bank_rows if r["origin"] == "Offers.processing" and r.get("source_ref") in ("Offers!2", "Offers!3")]
    # after collapse of processing keys, source_ref is min row of key — rows 2 and 3 should be different keys
    pnb_all = [
        r
        for r in bank_rows
        if r["origin"] == "Offers.processing" and bank_key(r.get("bank_name")).startswith("punjab national")
    ]
    pcts = {r.get("percentage") for r in pnb_all}
    asserts.check("spot_pnb_proc", 0 in pcts or 0.0 in pcts, f"pcts={pcts}")
    asserts.check("spot_pnb_proc_35", any(abs(float(p) - 0.0035) < 1e-12 for p in pcts if p is not None), f"pcts={pcts}")

    # DCB slab
    dcb = [r for r in bank_rows if r["origin"] == "Slab_Table"]
    asserts.check("spot_dcb_count", len(dcb) == 76, str(len(dcb)))
    first = sorted(dcb, key=lambda r: (float(r["slab_from"]) if r.get("slab_from") is not None else 0))[0]
    asserts.check(
        "spot_dcb_first",
        first.get("slab_from") == 0 and first.get("slab_to") == 3000 and first.get("fixed_amount") == 150,
        str(first),
    )

    # Canara overdue bands
    canara_od = [
        r
        for r in bank_rows
        if r["origin"] == "Offers.overdue" and "canara" in bank_key(r.get("bank_name"))
    ]
    asserts.check("spot_canara_od", len(canara_od) >= 1, str(len(canara_od)))

    # Govt unit (MODT first slab 0.1% points → 0.001 fraction; LA Stamp Duty intentionally omitted)
    modt = next((r for r in govt_out if r.get("charge_name") == "MODT Stamp Duty"), None)
    asserts.check(
        "govt_pct_0.001",
        modt is not None and abs(float(modt.get("percentage") or 0) - 0.001) < 1e-12,
        str(modt),
    )
    cersai = next((r for r in govt_out if r.get("charge_name") and "CERSAI" in str(r.get("charge_name"))), None)
    if cersai and not blank(cersai.get("flat_amount_inr")):
        asserts.check("cersai_flat", float(cersai.get("flat_amount_inr")) == 50, str(cersai.get("flat_amount_inr")))

    # Random 20 Other charges cell equality
    import random

    random.seed(42)
    sample = random.sample(oc_rows_src, min(20, len(oc_rows_src)))
    src_by_ref = {f"Other charges!{r}": d for r, d in oc_rows_src}
    for rnum, d in sample:
        out = find_oc(f"Other charges!{rnum}")
        ok = out is not None
        if ok:
            for h in OTHER_57:
                if h == "purpose" and d.get("purpose") == "Home Loan":
                    if out.get("purpose") != "Regular Home Loan":
                        ok = False
                    continue
                if out.get(h) != d.get(h):
                    # notes may have appended purpose_normalized
                    if h in ("note_1", "note_2") and d.get("purpose") == "Home Loan":
                        continue
                    ok = False
                    break
        asserts.check(f"spot_oc_sample_{rnum}", ok, "")

    # Query smoke
    axis_sal = [
        o
        for o in offers_out
        if "axis" in bank_key(o.get("bank_name"))
        and o.get("occupation") == "Salaried"
        and o.get("rate_type") == "Floating"
    ]
    # CIBIL 820 match scored bands
    axis_820 = []
    for o in axis_sal:
        if o.get("cibil_score_status") == "Not_Used":
            axis_820.append(o)
        elif o.get("cibil_score_status") == "Scored":
            try:
                if float(o["cibil_band_score_min"]) <= 820 <= float(o["cibil_band_score_max"]):
                    axis_820.append(o)
            except (TypeError, ValueError, KeyError):
                pass
    asserts.check("smoke_axis_offers", len(axis_820) >= 1, str(len(axis_820)))

    axis_key = bank_key("Axis Bank")
    axis_proc = [
        r
        for r in bank_rows
        if r.get("bank_key") == axis_key and r.get("charge_name") == "Processing fee"
    ]
    asserts.check("smoke_axis_proc", len(axis_proc) >= 1, str(len(axis_proc)))
    asserts.check(
        "smoke_join",
        sum(1 for r in bank_rows if r.get("bank_key") == axis_key) > 0,
        "",
    )
    before_axis = [
        r
        for r in bank_rows
        if r.get("bank_key") == axis_key and r.get("when_it_matters") == "Before offer"
    ]
    asserts.check("smoke_before", len(before_axis) >= 1, str(len(before_axis)))

    # Axis repricing group 4 rows
    if slabs[0]:
        gid = slabs[0]["charge_group_id"]
        grows = [r for r in bank_rows if r["charge_group_id"] == gid]
        asserts.check("smoke_repricing_4", len(grows) == 4, str(len(grows)))

    float_prepay = [
        r
        for r in bank_rows
        if r.get("charge_name") == "Prepayment charges" and r.get("rate_type") == "Floating"
    ]
    zeroish = sum(1 for r in float_prepay if r.get("fixed_amount") == 0)
    asserts.check("smoke_float_prepay", len(float_prepay) == 0 or zeroish / max(len(float_prepay), 1) >= 0.5, f"{zeroish}/{len(float_prepay)}")

    asserts.check("smoke_dcb_76", len(dcb) == 76, "")

    # Source unchanged
    asserts.check("source_mtime", SOURCE.stat().st_mtime == src_stat.st_mtime, "")
    asserts.check("source_size", SOURCE.stat().st_size == src_stat.st_size, "")
    asserts.check("source_sha_end", sha256_file(SOURCE) == LOCKED_SHA256, "")

    wb_src.close()

    # Write outputs only if pass? Plan: abort if fail. Still write VALIDATION always.
    # Write xlsx only on full pass.
    when_counts = Counter(r.get("when_it_matters") for r in bank_rows)
    report.update(
        {
            "digest": digest,
            "origin_counts": dict(origin_counters),
            "when_counts": dict(when_counts),
            "nr_names": nr_names,
            "govt_pct_log": govt_pct_log,
            "processing_ambiguous": processing_ambiguous,
            "collapse_appendix": collapse_appendix[:50],
            "collapse_total": len(collapse_appendix),
            "fixed_blank_premium": sorted(FIXED_BLANK_PREMIUM_ROWS),
            "slab_deferred": slab_deferred_rows,
        }
    )

    if asserts.ok():
        OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        write_sheet(wb_out, "Offers", OFFERS_OUT_HEADERS, offers_out)
        write_sheet(wb_out, "Bank_charges", BANK_HEADERS, bank_rows)
        write_sheet(wb_out, "Government_charges", GOVT_HEADERS, govt_out)
        asserts.check("sheet_count", len(wb_out.sheetnames) == 3, str(wb_out.sheetnames))
        wb_out.save(OUT_XLSX)
        print(f"Wrote {OUT_XLSX}")
    else:
        print("Conversion FAILED — xlsx not written.", file=sys.stderr)
        for f in asserts.failures[:30]:
            print(f, file=sys.stderr)
        if len(asserts.failures) > 30:
            print(f"... and {len(asserts.failures) - 30} more", file=sys.stderr)

    _write_validation(
        asserts,
        report,
        digest,
        src_stat,
        offers_out if asserts.ok() else offers_out,
        bank_rows,
        govt_out,
        {
            "proc": proc_groups,
            "od": od_groups,
            "pre": pre_groups,
        },
        abort=not asserts.ok(),
    )

    return 0 if asserts.ok() else 1


def _write_validation(
    asserts: AssertCollector,
    report: dict[str, Any],
    digest: str,
    src_stat,
    offers_out,
    bank_rows,
    govt_out,
    groups,
    abort: bool,
) -> None:
    OUT_VAL.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append("# HOME_LOANS_COMPARE_v1 VALIDATION")
    lines.append("")
    lines.append(f"**Result:** {'PASS' if not abort and asserts.ok() else 'FAIL'}")
    lines.append("")
    lines.append(f"- Source SHA-256: `{digest}`")
    lines.append(f"- Locked SHA-256: `{LOCKED_SHA256}`")
    lines.append(f"- Source mtime/size at start: {src_stat.st_mtime} / {src_stat.st_size}")
    lines.append("")
    lines.append("## Failures")
    if asserts.failures:
        for f in asserts.failures:
            lines.append(f"- {f}")
    else:
        lines.append("- (none)")
    lines.append("")
    lines.append("## Locked counts")
    if bank_rows is not None:
        oc = Counter(r.get("origin") for r in bank_rows)
        lines.append(f"- Offers: {len(offers_out) if offers_out else 'n/a'} (expect {LOCKED['offers']})")
        lines.append(f"- Other charges: {oc.get('Other charges', 0)} (expect {LOCKED['other_charges']})")
        lines.append(f"- Processing: {oc.get('Offers.processing', 0)} (expect {LOCKED['processing']})")
        lines.append(f"- Overdue: {oc.get('Offers.overdue', 0)} (expect {LOCKED['overdue']})")
        lines.append(f"- Slab_Table: {oc.get('Slab_Table', 0)} (expect {LOCKED['slab']})")
        lines.append(f"- Prepayment: {oc.get('Offers.prepayment', 0)} (expect {LOCKED['prepayment']})")
        lines.append(f"- Government: {len(govt_out) if govt_out else 'n/a'} (expect {LOCKED['government']})")
        lines.append(f"- when_it_matters: {report.get('when_counts')}")
        lines.append(f"- NEEDS_REVIEW names: {report.get('nr_names')}")
    lines.append("")
    lines.append("## JSON readiness")
    lines.append("- percentage_unit=fraction on Bank_charges and Government_charges")
    lines.append("- Government Percentage converted /100 (percent points → fraction)")
    for g in report.get("govt_pct_log", [])[:10]:
        lines.append(f"  - {g}")
    lines.append("- Purpose Home Loan → Regular Home Loan + note")
    lines.append("- Dates as YYYY-MM-DD text")
    lines.append("")
    lines.append("## Confirmations")
    lines.append("- One fee writing pattern: all fees are Bank_charges rows")
    lines.append("- No processing_fee_* / prepayment_* parallel amount columns in headers")
    lines.append("- product scripts can ignore `origin` (audit only)")
    lines.append("- Other charges charge_name text unchanged")
    lines.append("- Safe Custody / Other Services / Other miscellaneous approvals → After offer (curator)")
    lines.append("- Loan Agreement Stamp Duty omitted from Government_charges (curator deleted)")
    lines.append("- req_amount_not_discriminative=true (every Offers row 1..1e9)")
    lines.append("- Fixed blank premium treated as 0 in ROI assert for rows: " + str(report.get("fixed_blank_premium")))
    lines.append("")
    lines.append("## bank_key join coverage")
    if offers_out and bank_rows:
        ok = {o["bank_key"] for o in offers_out}
        bk = {r["bank_key"] for r in bank_rows if r.get("bank_key")}
        lines.append(f"- Offers distinct bank_key: {len(ok)}")
        lines.append(f"- Bank_charges distinct bank_key: {len(bk)}")
        lines.append(f"- Offers keys with ≥1 Bank_charges row: {len(ok & bk)}")
        lines.append(f"- Offers keys with no Bank_charges: {sorted(ok - bk)[:20]}")
    lines.append("")
    lines.append("## PROCESSING_MATCH_AMBIGUOUS")
    lines.append(f"- groups: {len(report.get('processing_ambiguous', []))}")
    for a in report.get("processing_ambiguous", [])[:20]:
        lines.append(f"  - {a}")
    lines.append("")
    lines.append("## Collapse appendix (sample)")
    lines.append(f"- Total collapsed groups with >1 source row: {report.get('collapse_total')}")
    for line in report.get("collapse_appendix", [])[:30]:
        lines.append(f"- {line}")
    lines.append("")
    lines.append("## Slab_Table deferred Offers rows")
    lines.append(str(report.get("slab_deferred")))
    lines.append("")
    lines.append("## CIBIL status engine notes")
    lines.append("- Not_Used — do not filter on customer CIBIL")
    lines.append("- No_Score — match only no-score/NTC input; never numeric 720")
    lines.append("- Thin_File — match thin-file input only; not normal 300–900")
    lines.append("- Scored — customer score in [min, max]")
    lines.append("")
    OUT_VAL.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_VAL}")


if __name__ == "__main__":
    sys.exit(main())

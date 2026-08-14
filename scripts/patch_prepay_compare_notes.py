#!/usr/bin/env python3
"""Add clarifying notes on prepayment-statement document fees in Bank_charges."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_COMPARE = ROOT / "data" / "HOME_LOANS_COMPARE_v1.xlsx"

STATEMENT_NOTE = (
    "Document fee for prepayment statement printout; not a prepayment penalty."
)

TARGETS = {
    ("City Union Bank", "Prepayment Statement Charges"),
    ("ICICI Bank", "Charges for Prepayment Statement"),
}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--compare", type=Path, default=DEFAULT_COMPARE)
    args = ap.parse_args()
    if not args.compare.exists():
        print(f"Missing compare file: {args.compare}", file=sys.stderr)
        return 1

    wb = load_workbook(args.compare)
    ws = wb["Bank_charges"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    patched = 0
    for row in ws.iter_rows(min_row=2):
        bank = row[idx["bank_name"]].value
        name = row[idx["charge_name"]].value
        key = (bank, name)
        if key not in TARGETS:
            continue
        note_cell = row[idx["note_1"]]
        if note_cell.value != STATEMENT_NOTE:
            note_cell.value = STATEMENT_NOTE
            patched += 1

    wb.save(args.compare)
    wb.close()
    print(f"Patched statement-fee notes: {patched}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

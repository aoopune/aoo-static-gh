#!/usr/bin/env python3
"""Convert the corrected APF workbook into versioned client-side JSON."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = (
    ROOT.parent
    / "etc"
    / "HOME LOANS - OFFERS - APF with localities - CORRECTED Jul21.xlsx"
)
DEFAULT_OUTPUT = ROOT / "data" / "apf-home-loan-projects.json"
SHEET_NAME = "APF as per Banks"
EXPECTED_HEADERS = [
    "bank_name",
    "Developer_name",
    "project_name",
    "Area / Locality",
    "Full Address",
    "City",
    "rera_no.",
    "apf_code",
    "Branch / Contact",
    "area_name",
    "Official Address",
    "Lookup Score",
    "Lookup Sources",
]

RERA_PATTERN = re.compile(r"\bP\d{11}\b", re.IGNORECASE)
URL_PATTERN = re.compile(r"^https?://", re.IGNORECASE)

# Verified project-site locations for rows corrected in the copied workbook.
LOCATION_OVERRIDES = {
    4582: ("Maharashtra", "Satara"),
    4623: ("Maharashtra", "Washim"),
    4644: ("Maharashtra", "Kolhapur"),
    4673: ("Maharashtra", "Satara"),
    4728: ("Maharashtra", "Satara"),
    5977: ("Goa", "North Goa"),
    6004: ("Maharashtra", "Satara"),
    6411: ("Maharashtra", "Satara"),
    7317: ("Maharashtra", "Solapur"),
    7389: ("Maharashtra", "Satara"),
    7579: ("Maharashtra", "Solapur"),
    7590: ("Maharashtra", "Thane"),
    8080: ("Maharashtra", "Kolhapur"),
    8465: ("Maharashtra", "Solapur"),
}

LOCATION_REVIEW_ROWS = {4562, 6474, 6488}
RERA_REVIEW_ROWS = {6175, 6923, 7008, 7322, 7946, 8289, 8335, 8369, 8739}


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def normalize(value: Any) -> str:
    raw = text(value) or ""
    ascii_text = unicodedata.normalize("NFKD", raw)
    ascii_text = "".join(char for char in ascii_text if not unicodedata.combining(char))
    return " ".join(re.sub(r"[^a-z0-9]+", " ", ascii_text.lower()).split())


def split_sources(value: Any) -> list[str]:
    raw = text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split("|") if URL_PATTERN.match(part.strip())]


def extract_rera_numbers(value: Any) -> list[str]:
    raw = text(value) or ""
    return list(dict.fromkeys(match.upper() for match in RERA_PATTERN.findall(raw)))


def serializable_raw(values: list[Any]) -> list[Any]:
    result: list[Any] = []
    for value in values:
        if value is None or isinstance(value, (str, int, float, bool)):
            result.append(value)
        else:
            result.append(str(value))
    return result


def build_dataset(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, read_only=True, data_only=False)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Missing sheet: {SHEET_NAME}")

    worksheet = workbook[SHEET_NAME]
    headers = [text(worksheet.cell(1, column).value) for column in range(1, 14)]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected A1:M1 headers: {headers!r}")

    rows: list[tuple[int, list[Any]]] = []
    for row_number, cells in enumerate(
        worksheet.iter_rows(min_row=2, max_col=13, values_only=True), start=2
    ):
        values = serializable_raw(list(cells))
        if not any(value not in (None, "") for value in values):
            continue
        rows.append((row_number, values))

    first_seen: dict[str, int] = {}
    duplicate_rows = 0
    records: list[dict[str, Any]] = []

    for row_number, values in rows:
        (
            bank_name,
            developer_name,
            project_name,
            area_locality,
            full_address,
            city,
            rera_raw,
            apf_raw,
            branch_contact,
            area_name,
            official_address,
            lookup_score,
            lookup_sources_raw,
        ) = values

        fingerprint = json.dumps(values, ensure_ascii=False, separators=(",", ":"))
        duplicate_of = first_seen.get(fingerprint)
        if duplicate_of is None:
            first_seen[fingerprint] = row_number
        else:
            duplicate_rows += 1

        state_name, district_name = LOCATION_OVERRIDES.get(
            row_number, ("Maharashtra", "Pune")
        )
        location_status = "verified_override" if row_number in LOCATION_OVERRIDES else "default_pune"
        if row_number in LOCATION_REVIEW_ROWS:
            location_status = "review"

        rera_numbers = extract_rera_numbers(rera_raw)
        rera_text = text(rera_raw)
        sources = split_sources(lookup_sources_raw)

        record = {
            "id": f"apf-row-{row_number:05d}",
            "source_row": row_number,
            "bank_name": text(bank_name),
            "developer_name": text(developer_name),
            "project_name": text(project_name),
            "area_locality": text(area_locality),
            "full_address": text(full_address),
            "city": text(city),
            "state_name": state_name,
            "district_name": district_name,
            "area_name": text(area_name),
            "rera_no_raw": rera_text,
            "rera_numbers": rera_numbers,
            "apf_code": text(apf_raw),
            "branch_contact": text(branch_contact),
            "official_address": text(official_address),
            "locality_lookup_score": lookup_score,
            "locality_lookup_sources_raw": text(lookup_sources_raw),
            "locality_lookup_sources": sources,
            "search": {
                "developer_name": normalize(developer_name),
                "project_name": normalize(project_name),
                "area_name": normalize(area_name),
            },
            "quality": {
                "exact_duplicate": duplicate_of is not None,
                "duplicate_of_source_row": duplicate_of,
                "location_status": location_status,
                "rera_parse_warning": bool(rera_text and not rera_numbers)
                or row_number in RERA_REVIEW_ROWS,
                "source_url_warning": bool(text(lookup_sources_raw) and not sources),
            },
        }
        records.append(record)

    source_stat = input_path.stat()
    return {
        "schema_version": 1,
        "metadata": {
            "product": "home_loan",
            "source_workbook": input_path.name,
            "source_sheet": SHEET_NAME,
            "source_sha256": file_sha256(input_path),
            "source_modified_at": datetime.fromtimestamp(
                source_stat.st_mtime, timezone.utc
            ).isoformat(),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "excluded_sheets": ["Out of Pune - removed"],
            "record_count": len(records),
            "exact_duplicate_record_count": duplicate_rows,
            "default_location": {
                "state_name": "Maharashtra",
                "district_name": "Pune",
            },
        },
        "records": records,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    input_path = args.input.expanduser().resolve()
    output_path = args.output.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    dataset = build_dataset(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(dataset, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(
        f"Wrote {dataset['metadata']['record_count']} records to {output_path} "
        f"({output_path.stat().st_size:,} bytes)"
    )


if __name__ == "__main__":
    main()

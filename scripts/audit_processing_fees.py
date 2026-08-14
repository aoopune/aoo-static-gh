#!/usr/bin/env python3
"""Legacy audit: source Offers → master Bank_charges + reference sheet.

For post-cleanup verification against your human sheet only, use:
  scripts/audit_processing_fees_human.py
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "HOME LOANS - OFFERS (5).xlsx"
MASTER = ROOT / "data" / "HOME_LOANS_COMPARE_v1.xlsx"
REF = ROOT / "data" / "Home Loans.xlsx"
OUT_DIR = ROOT / "data" / "Charges" / "_audit"
OUT_SUMMARY = OUT_DIR / "PROCESSING_FEES_AUDIT.md"
OUT_JSON = OUT_DIR / "processing_fees_audit.json"

NAINITAL_DROP = {1363, 1379, 1395, 1397, 1399, 1401, 1403}

PROC_KEY = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type", "occupation",
    "borrower_category", "cibil_band_applicable", "cibil_band_score_min",
    "cibil_band_score_max", "loan_amount_band_applicable", "loan_amount_min",
    "loan_amount_max", "tenure_band_applicable", "tenure_months_min",
    "tenure_months_max", "processing_fee_calc_type", "processing_fee_percent",
    "processing_fee_flat_amount", "processing_fee_min_amount", "processing_fee_max_amount",
    "processing_fee_valid_from", "processing_fee_valid_till",
]
PROC_MATCH_NO_AMT = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type", "occupation",
    "borrower_category", "cibil_band_applicable", "cibil_band_score_min",
    "cibil_band_score_max", "loan_amount_band_applicable", "loan_amount_min",
    "loan_amount_max", "tenure_band_applicable", "tenure_months_min",
    "tenure_months_max", "processing_fee_valid_from", "processing_fee_valid_till",
]

MASTER_COMPARE_FIELDS = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type", "occupation",
    "borrower_category", "cibil_band_applicable", "cibil_band_score_min",
    "cibil_band_score_max", "loan_amount_band_applicable", "loan_amount_min",
    "loan_amount_max", "tenure_band_applicable", "tenure_months_min",
    "tenure_months_max", "charge_type", "percentage", "fixed_amount",
    "charge_min", "charge_max", "valid_from", "valid_till",
]

REF_NAME_MAP = {
    "central bank of india": "Central Bank of India",
    "bank of maharastra": "Bank of Maharashtra",
    "federal bank": "Federal Bank",
    "idfc first bank": "IDFC FIRST Bank",
    "indusind bank": "IndusInd Bank",
    "jammu & kashmir bank": "Jammu and Kashmir Bank",
    "karur vyasa bank": "Karur Vysya Bank",
    "naintal bank": "Nainital Bank",
}


def blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def bank_key(name: Any) -> str:
    s = "" if name is None else str(name)
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("&", "and")
    return s


def iso_date(v: Any) -> Any:
    if blank(v):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            return s
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s[:10], fmt).date().isoformat()
            except ValueError:
                continue
        return s
    return str(v)


def num(v: Any) -> float | None:
    if blank(v):
        return None
    return float(v)


def row_dict(headers: list[str], values: tuple | list) -> dict[str, Any]:
    vals = list(values)
    if len(vals) < len(headers):
        vals += [None] * (len(headers) - len(vals))
    return {h: vals[i] for i, h in enumerate(headers)}


def key_tuple(d: dict[str, Any], fields: list[str]) -> tuple:
    return tuple(d.get(f) for f in fields)


def norm_ref_bank(name: str) -> str:
    k = bank_key(name)
    return REF_NAME_MAP.get(k, str(name).strip())


def source_to_master_row(d: dict[str, Any]) -> dict[str, Any]:
    calc = d.get("processing_fee_calc_type")
    row = {f: d.get(f) for f in MASTER_COMPARE_FIELDS}
    if calc == "Percentage":
        row["charge_type"] = "Percentage"
        row["percentage"] = d.get("processing_fee_percent")
        row["fixed_amount"] = None
    elif calc == "Flat":
        row["charge_type"] = "Fixed Amount"
        row["fixed_amount"] = d.get("processing_fee_flat_amount")
        row["percentage"] = None
    else:
        row["charge_type"] = calc
        row["percentage"] = d.get("processing_fee_percent")
        row["fixed_amount"] = d.get("processing_fee_flat_amount")
    row["charge_min"] = d.get("processing_fee_min_amount")
    row["charge_max"] = d.get("processing_fee_max_amount")
    row["valid_from"] = iso_date(d.get("processing_fee_valid_from"))
    row["valid_till"] = iso_date(d.get("processing_fee_valid_till"))
    return row


def master_row_norm(r: dict[str, Any]) -> dict[str, Any]:
    return {
        f: (iso_date(r.get(f)) if f in ("valid_from", "valid_till") else r.get(f))
        for f in MASTER_COMPARE_FIELDS
    }


def rows_equal(a: dict[str, Any], b: dict[str, Any]) -> tuple[bool, list[str]]:
    diffs: list[str] = []
    for f in MASTER_COMPARE_FIELDS:
        av, bv = a.get(f), b.get(f)
        if f in ("percentage", "fixed_amount", "charge_min", "charge_max",
                 "loan_amount_min", "loan_amount_max", "tenure_months_min",
                 "tenure_months_max", "cibil_band_score_min", "cibil_band_score_max"):
            an, bn = num(av), num(bv)
            if an is None and bn is None:
                continue
            if an is None or bn is None or abs(an - bn) > 1e-9:
                diffs.append(f"{f}: {av!r} vs {bv!r}")
        elif iso_date(av) if f in ("valid_from", "valid_till") else av != bv:
            if f in ("valid_from", "valid_till"):
                if iso_date(av) != iso_date(bv):
                    diffs.append(f"{f}: {av!r} vs {bv!r}")
            else:
                diffs.append(f"{f}: {av!r} vs {bv!r}")
    return not diffs, diffs


def fingerprint(row: dict[str, Any]) -> tuple:
    return (
        row.get("charge_type"),
        num(row.get("percentage")),
        num(row.get("fixed_amount")),
        num(row.get("charge_min")),
        num(row.get("charge_max")),
        num(row.get("loan_amount_min")),
        num(row.get("loan_amount_max")),
        row.get("cibil_band_applicable"),
        num(row.get("cibil_band_score_min")),
        num(row.get("cibil_band_score_max")),
    )


def parse_ref_text(text: str) -> dict[str, Any]:
    t = text or ""
    out: dict[str, Any] = {"raw": t}
    pct = re.findall(r"(\d+(?:\.\d+)?)\s*%", t.replace(",", ""))
    if pct:
        out["percentages"] = sorted({float(x) / 100 for x in pct})
    mins = re.findall(r"min\.?\s*[₹:]?\s*(\d[\d,]*)", t, re.I)
    maxs = re.findall(r"max\.?\s*[₹:]?\s*(\d[\d,]*)", t, re.I)
    if mins:
        out["mins"] = sorted({int(x.replace(",", "")) for x in mins})
    if maxs:
        out["maxs"] = sorted({int(x.replace(",", "")) for x in maxs})
    if re.search(r"\bnil\b", t, re.I):
        out["has_nil"] = True
    if re.search(r"whichever is higher|which ever is higher", t, re.I):
        out["whichever_higher"] = True
    return out


def load_source_processing() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Offers"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    filled: list[tuple[int, dict[str, Any]]] = []
    for rnum, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        d = row_dict(headers, raw)
        if all(blank(d.get(h)) for h in headers):
            continue
        if rnum in NAINITAL_DROP:
            continue
        filled.append((rnum, d))

    proc_groups: dict[tuple, list[int]] = defaultdict(list)
    for r, d in filled:
        proc_groups[key_tuple(d, PROC_KEY)].append(r)

    ambig_match: dict[tuple, list[tuple]] = defaultdict(list)
    for k in proc_groups:
        ambig_match[tuple(k[i] for i, f in enumerate(PROC_KEY) if f in PROC_MATCH_NO_AMT)].append(k)

    ambiguous: list[dict[str, Any]] = []
    ambig_keys: set[tuple] = set()
    for mk, keys in ambig_match.items():
        if len(keys) > 1:
            for k in keys:
                ambig_keys.add(k)
            ambiguous.append({"match": mk, "keys": len(keys), "source_rows": [min(proc_groups[k]) for k in keys]})

    extracted: list[dict[str, Any]] = []
    for k, rnums in sorted(proc_groups.items(), key=lambda x: min(x[1])):
        src_row = min(rnums)
        d = next(d for r, d in filled if r == src_row)
        row = source_to_master_row(d)
        row["_source_row"] = src_row
        row["_contributor_rows"] = sorted(rnums)
        row["_ambiguous"] = k in ambig_keys
        row["bank_key"] = bank_key(d["bank_name"])
        extracted.append(row)
    wb.close()
    return extracted, ambiguous


def load_master_processing() -> list[dict[str, Any]]:
    wb = load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb["Bank_charges"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        d = row_dict(headers, raw)
        if d.get("origin") != "Offers.processing":
            continue
        r = master_row_norm(d)
        r["charge_row_id"] = d.get("charge_row_id")
        r["source_ref"] = d.get("source_ref")
        r["bank_key"] = d.get("bank_key")
        r["note_1"] = d.get("note_1")
        r["note_2"] = d.get("note_2")
        rows.append(r)
    wb.close()
    return rows


def load_reference_sheet() -> dict[str, dict[str, Any]]:
    wb = load_workbook(REF, read_only=True, data_only=True)
    ws = wb["Processing fees"]
    out: dict[str, dict[str, Any]] = {}
    for raw in ws.iter_rows(min_row=3, values_only=True):
        if blank(raw[0]):
            continue
        bank = norm_ref_bank(str(raw[0]))
        out[bank] = {
            "bank_name": bank,
            "ref_text": raw[1] if len(raw) > 1 else None,
            "validity_note": raw[2] if len(raw) > 2 else None,
            "url": raw[3] if len(raw) > 3 else None,
            "parsed": parse_ref_text(str(raw[1] or "")),
        }
    wb.close()
    return out


def find_exact_duplicates(rows: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    groups: dict[tuple, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        k = tuple(r.get(f) for f in MASTER_COMPARE_FIELDS)
        groups[k].append(r)
    return [g for g in groups.values() if len(g) > 1]


def find_redundant_dimensions(rows: list[dict[str, Any]]) -> list[str]:
    notes: list[str] = []
    fps = Counter(fingerprint(r) for r in rows)
    dup_fps = {fp: c for fp, c in fps.items() if c > 1}
    if dup_fps:
        notes.append(f"{len(dup_fps)} amount fingerprint(s) repeat across {sum(dup_fps.values())} rows — check if filter dimensions are redundant")
    # same amount, all filter fields identical except one always-Any
    return notes


def cross_check_ref(bank: str, rows: list[dict[str, Any]], ref: dict[str, Any] | None) -> list[str]:
    issues: list[str] = []
    if not ref:
        issues.append("Bank missing from Processing fees reference sheet (name may differ — check spelling)")
        return issues
    parsed = ref["parsed"]
    pcts = {round(x, 6) for x in (num(r.get("percentage")) for r in rows if r.get("charge_type") == "Percentage") if x is not None}
    mins = {int(num(r.get("charge_min"))) for r in rows if not blank(r.get("charge_min"))}
    maxs = {int(num(r.get("charge_max"))) for r in rows if not blank(r.get("charge_max"))}
    has_zero = any(num(r.get("percentage")) == 0 or num(r.get("fixed_amount")) == 0 for r in rows)

    if parsed.get("percentages"):
        for rp in parsed["percentages"]:
            rp6 = round(rp, 6)
            if rp6 not in pcts and not (rp == 0 and has_zero):
                if not any(abs(rp - p) < 1e-6 for p in pcts):
                    # ignore obvious GST / mis-parse (e.g. BOB "50%" meaning 0.50%)
                    if rp >= 0.1 and any(abs(rp / 100 - p) < 1e-6 for p in pcts):
                        continue
                    issues.append(f"Reference mentions {rp*100:.2f}% but master distinct percentages are {[round(x*100,4) if x else 0 for x in sorted(pcts)]}")
    if parsed.get("mins"):
        for rm in parsed["mins"]:
            if rm not in mins:
                issues.append(f"Reference min ₹{rm:,} not in master mins {sorted(mins)}")
    if parsed.get("maxs"):
        for rx in parsed["maxs"]:
            if rx not in maxs:
                issues.append(f"Reference max ₹{rx:,} not in master maxes {sorted(maxs)}")
    if parsed.get("has_nil") and not has_zero:
        issues.append("Reference mentions NIL but no zero-fee rows in master")
    return issues


def audit_bank(
    bank: str,
    source_rows: list[dict[str, Any]],
    master_rows: list[dict[str, Any]],
    ref: dict[str, Any] | None,
    global_ambiguous: list[dict[str, Any]],
) -> dict[str, Any]:
    src_by_key = {tuple(r.get(f) for f in MASTER_COMPARE_FIELDS): r for r in source_rows}
    mst_by_key = {tuple(r.get(f) for f in MASTER_COMPARE_FIELDS): r for r in master_rows}

    missing_in_master = []
    for k, r in src_by_key.items():
        if k not in mst_by_key:
            missing_in_master.append(r)

    extra_in_master = []
    for k, r in mst_by_key.items():
        if k not in src_by_key:
            extra_in_master.append(r)

    value_mismatches = []
    for k in set(src_by_key) & set(mst_by_key):
        ok, diffs = rows_equal(src_by_key[k], mst_by_key[k])
        if not ok:
            value_mismatches.append({"source_row": src_by_key[k].get("_source_row"), "diffs": diffs})

    exact_dups = find_exact_duplicates(master_rows)
    ref_issues = cross_check_ref(bank, master_rows, ref)

    bank_ambig = [
        a for a in global_ambiguous
        if any(bank_key(r.get("bank_name")) == bank_key(bank) for r in source_rows
               for sr in a["source_rows"])
    ]

    redundant_notes = find_redundant_dimensions(master_rows)

    # Non-processing charge names wrongly tagged?
    wrong_names = [r for r in master_rows if r.get("charge_name") not in (None, "Processing fee")]

    verdict = "PASS"
    if missing_in_master or extra_in_master or value_mismatches or exact_dups or ref_issues:
        verdict = "FAIL"
    elif bank_ambig:
        verdict = "REVIEW"

    return {
        "bank": bank,
        "verdict": verdict,
        "source_rows": len(source_rows),
        "master_rows": len(master_rows),
        "missing_in_master": len(missing_in_master),
        "extra_in_master": len(extra_in_master),
        "value_mismatches": len(value_mismatches),
        "exact_duplicates": len(exact_dups),
        "duplicate_row_ids": [[r.get("charge_row_id") for r in g] for g in exact_dups],
        "source_ambiguous_groups": len(bank_ambig),
        "ref_issues": ref_issues,
        "redundant_notes": redundant_notes,
        "wrong_charge_names": len(wrong_names),
        "fingerprints": {fmt_fp(fp): cnt for fp, cnt in Counter(fingerprint(r) for r in master_rows).items()},
        "details": {
            "missing_sample": missing_in_master[:5],
            "extra_sample": extra_in_master[:5],
            "mismatch_sample": value_mismatches[:5],
            "ref_text": (ref or {}).get("ref_text"),
        },
    }


def fmt_fp(fp: tuple) -> str:
    ct, pct, flat, cmin, cmax, lmin, lmax, cibil, csmin, csmax = fp
    parts = []
    if ct == "Percentage" and pct is not None:
        parts.append(f"{pct*100:.2g}%")
    elif ct == "Fixed Amount" and flat is not None:
        parts.append(f"₹{flat:,.0f} flat")
    if cmin is not None:
        parts.append(f"min ₹{cmin:,.0f}")
    if cmax is not None:
        parts.append(f"max ₹{cmax:,.0f}")
    if lmin is not None or lmax is not None:
        parts.append(f"loan {lmin or '—'}–{lmax or '—'}")
    if cibil == "Yes" and (csmin is not None or csmax is not None):
        parts.append(f"CIBIL {csmin}–{csmax}")
    return "; ".join(parts) if parts else str(fp)


def render_markdown(results: list[dict[str, Any]], ambiguous: list[dict[str, Any]], totals: dict[str, Any]) -> str:
    lines = [
        "# Processing Fees Audit",
        "",
        "Compares structured source (`HOME LOANS - OFFERS (5).xlsx` → **Offers** sheet, same logic as converter)",
        "against master (`HOME_LOANS_COMPARE_v1.xlsx` → **Bank_charges**, `origin=Offers.processing`),",
        "with narrative cross-check against `Home Loans.xlsx` → **Processing fees** reference sheet.",
        "",
        "> **Note:** The **Processing fees** sheet in `Home Loans.xlsx` is a human summary (one row per bank).",
        "> Structured fee rows live in the Offers workbook and are exploded into many master rows by scheme, CIBIL band, loan amount, etc.",
        "",
        "## Overall",
        "",
        f"| Metric | Count |",
        f"|---|---|",
        f"| Banks | {totals['banks']} |",
        f"| Source processing keys | {totals['source_keys']} |",
        f"| Master processing rows | {totals['master_rows']} |",
        f"| Missing in master | {totals['missing']} |",
        f"| Extra in master | {totals['extra']} |",
        f"| Value mismatches | {totals['mismatches']} |",
        f"| Exact duplicate groups (master) | {totals['dup_groups']} |",
        f"| Source ambiguous groups (same match, different amounts) | {len(ambiguous)} |",
        f"| Banks PASS | {totals['pass']} |",
        f"| Banks FAIL | {totals['fail']} |",
        f"| Banks REVIEW | {totals['review']} |",
        "",
        "## Per-bank summary",
        "",
        "| Bank | Verdict | Source rows | Master rows | Missing | Extra | Mismatches | Duplicates | Ref issues |",
        "|---|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for r in sorted(results, key=lambda x: (0 if x["verdict"] == "FAIL" else 1 if x["verdict"] == "REVIEW" else 2, x["bank"])):
        lines.append(
            f"| {r['bank']} | {r['verdict']} | {r['source_rows']} | {r['master_rows']} | "
            f"{r['missing_in_master']} | {r['extra_in_master']} | {r['value_mismatches']} | "
            f"{r['exact_duplicates']} | {len(r['ref_issues'])} |"
        )

    lines.extend(["", "## Bank details", ""])
    for r in sorted(results, key=lambda x: x["bank"]):
        lines.append(f"### {r['bank']} — {r['verdict']}")
        lines.append(f"- Source keys: {r['source_rows']} → Master rows: {r['master_rows']}")
        if r["missing_in_master"]:
            lines.append(f"- **Missing in master:** {r['missing_in_master']}")
            for s in r["details"]["missing_sample"]:
                lines.append(f"  - Offers!{s.get('_source_row')}: {fmt_fp(fingerprint(s))}")
        if r["extra_in_master"]:
            lines.append(f"- **Extra in master:** {r['extra_in_master']}")
            for s in r["details"]["extra_sample"]:
                lines.append(f"  - {s.get('charge_row_id')} ({s.get('source_ref')})")
        if r["value_mismatches"]:
            lines.append(f"- **Value mismatches:** {r['value_mismatches']}")
            for m in r["details"]["mismatch_sample"]:
                lines.append(f"  - Offers!{m['source_row']}: {', '.join(m['diffs'][:4])}")
        if r["exact_duplicates"]:
            lines.append(f"- **Exact duplicate groups:** {r['exact_duplicates']}")
            for ids in r["duplicate_row_ids"][:3]:
                lines.append(f"  - {', '.join(ids)}")
        if r["source_ambiguous_groups"]:
            lines.append(f"- **Source ambiguity:** {r['source_ambiguous_groups']} group(s) — same offer match dimensions but different fee amounts in source Offers sheet")
        if r["ref_issues"]:
            lines.append("- **Reference sheet cross-check issues:**")
            for issue in r["ref_issues"]:
                lines.append(f"  - {issue}")
        elif r["details"].get("ref_text"):
            lines.append("- Reference sheet: no obvious numeric mismatch detected (exploded rows may still differ in granularity)")
        fps = r["fingerprints"]
        if fps:
            lines.append(f"- Distinct fee fingerprints in master ({len(fps)}):")
            for label, cnt in sorted(fps.items(), key=lambda x: -x[1])[:8]:
                lines.append(f"  - ×{cnt}: {label}")
        if r["redundant_notes"]:
            for n in r["redundant_notes"]:
                lines.append(f"- Note: {n}")
        lines.append("")

    if ambiguous:
        lines.extend(["## Source ambiguous groups (global)", ""])
        lines.append("Same bank/scheme/filter match but different processing amounts in Offers — flagged `SOURCE_AMBIGUOUS_PROCESSING_SAME_MATCH` in master.")
        for a in ambiguous[:30]:
            mk = str(a["match"])
            lines.append(f"- match={mk[:120]}… keys={a['keys']} rows={a['source_rows']}")
        if len(ambiguous) > 30:
            lines.append(f"- … and {len(ambiguous)-30} more")
        lines.append("")

    lines.extend([
        "## Redundant / non-processing data check",
        "",
        "- Processing rows should have `charge_name=Processing fee`, `origin=Offers.processing`, `when_it_matters=Before offer`.",
        "- Filter columns (scheme, CIBIL band, loan amount, tenure, occupation) are **intentional** — they are not duplicates unless two rows are identical on all dimensions.",
        "- Rows with `note_2` containing `SOURCE_AMBIGUOUS_PROCESSING_SAME_MATCH` need human pick of authoritative amount.",
        "- Processing fee rows must **not** appear in individual bank Charges Structured_Data sheets (by design).",
        "",
    ])
    return "\n".join(lines)


def main() -> int:
    source_rows, ambiguous = load_source_processing()
    master_rows = load_master_processing()
    ref_map = load_reference_sheet()

    by_bank_src: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in source_rows:
        by_bank_src[r["bank_name"]].append(r)
    by_bank_mst: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in master_rows:
        by_bank_mst[r["bank_name"]].append(r)

    all_banks = sorted(set(by_bank_src) | set(by_bank_mst) | set(ref_map))
    results = [
        audit_bank(b, by_bank_src.get(b, []), by_bank_mst.get(b, []), ref_map.get(b), ambiguous)
        for b in all_banks
    ]

    totals = {
        "banks": len(all_banks),
        "source_keys": len(source_rows),
        "master_rows": len(master_rows),
        "missing": sum(r["missing_in_master"] for r in results),
        "extra": sum(r["extra_in_master"] for r in results),
        "mismatches": sum(r["value_mismatches"] for r in results),
        "dup_groups": sum(r["exact_duplicates"] for r in results),
        "pass": sum(1 for r in results if r["verdict"] == "PASS"),
        "fail": sum(1 for r in results if r["verdict"] == "FAIL"),
        "review": sum(1 for r in results if r["verdict"] == "REVIEW"),
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    md = render_markdown(results, ambiguous, totals)
    OUT_SUMMARY.write_text(md, encoding="utf-8")
    ambig_json = [
        {"match": str(a["match"]), "keys": a["keys"], "source_rows": a["source_rows"]}
        for a in ambiguous
    ]
    OUT_JSON.write_text(
        json.dumps({"totals": totals, "ambiguous": ambig_json, "banks": results}, indent=2, default=str),
        encoding="utf-8",
    )

    print(md[:4000])
    print(f"\n… full report: {OUT_SUMMARY}")
    print(f"JSON: {OUT_JSON}")
    return 0 if totals["missing"] == 0 and totals["extra"] == 0 and totals["mismatches"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""
Surgical upsert: Fixed-rate prepayment rows only into HOME_LOANS_COMPARE_v1.xlsx Bank_charges.
Does not modify Offers, Government_charges, Floating prepay, or any other charges.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_STRUCTURED = ROOT / "data" / "fixed-prepay-structured.csv"
DEFAULT_COMPARE = ROOT / "data" / "HOME_LOANS_COMPARE_v1.xlsx"

NEW_COLS = [
    "months_from_event_min",
    "months_from_event_max",
    "months_from_event_basis",
    "percentage_applies_per",
    "source_url",
]

PREPAY_NAMES = {"Prepayment charges", "Prepayment charges (takeover)"}


def is_fixed_prepay(row: dict[str, Any]) -> bool:
    return (
        row.get("charge_name") in PREPAY_NAMES
        and str(row.get("rate_type") or "") == "Fixed"
    )


def blank_to_none(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def parse_number(v: Any) -> Any:
    v = blank_to_none(v)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return s


def read_structured(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8", newline="") as f:
        return [{k: blank_to_none(v) for k, v in row.items()} for row in csv.DictReader(f)]


def sheet_to_dicts(ws) -> tuple[list[str], list[dict[str, Any]]]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if any(h is None or not isinstance(h, str) for h in headers):
        raise SystemExit(f"Bad headers: {headers!r}")
    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        vals = list(raw)
        if len(vals) < len(headers):
            vals += [None] * (len(headers) - len(vals))
        elif len(vals) > len(headers):
            vals = vals[: len(headers)]
        if all(v is None or v == "" for v in vals):
            continue
        rows.append({headers[i]: vals[i] for i in range(len(headers))})
    return headers, rows


def max_pre_id(rows: list[dict[str, Any]]) -> int:
    mx = 0
    for r in rows:
        rid = str(r.get("charge_row_id") or "")
        m = re.match(r"CHG-PRE-(\d+)$", rid)
        if m:
            mx = max(mx, int(m.group(1)))
    return mx


def structured_to_bank_row(
    src: dict[str, Any], headers: list[str], next_id: int
) -> dict[str, Any]:
    row = {h: None for h in headers}
    rid = f"CHG-PRE-{next_id}"
    row["charge_row_id"] = rid
    row["charge_group_id"] = rid
    row["bank_key"] = src["bank_key"]
    row["bank_name"] = src["bank_name"]
    row["origin"] = "CSV.fixed_prepay"
    row["when_it_matters"] = "After offer"
    row["source_ref"] = src.get("csv_cell_ref")
    row["percentage_unit"] = "fraction"
    row["charge_name"] = src["charge_name"]
    row["facility_type"] = src["facility_type"]
    row["rate_type"] = "Fixed"
    row["purpose"] = "Regular Home Loan"
    row["has_slab_wise_charges"] = src.get("has_slab_wise_charges") or "No"
    row["slab_from"] = parse_number(src.get("slab_from"))
    row["slab_to"] = parse_number(src.get("slab_to"))
    row["slab_basis"] = src.get("slab_basis")
    if src.get("loan_amount_band_applicable"):
        row["loan_amount_band_applicable"] = src.get("loan_amount_band_applicable")
        row["loan_amount_min"] = parse_number(src.get("loan_amount_min"))
        row["loan_amount_max"] = parse_number(src.get("loan_amount_max"))
    status = src.get("charge_status")
    if status == "nil":
        row["charge_type"] = "Fixed Amount"
        row["fixed_amount"] = 0
        row["percentage"] = None
        row["percentage_base_value"] = None
        row["gst_applicable"] = None
        row["note_1"] = src.get("note_1") or "Prepayment nil (CSV NIL)"
    else:
        row["charge_type"] = "Percentage"
        row["fixed_amount"] = None
        row["percentage"] = parse_number(src.get("percentage"))
        row["percentage_base_value"] = src.get("percentage_base_value")
        row["gst_applicable"] = src.get("gst_applicable") or "Yes"
        row["note_1"] = src.get("note_1")
    if "months_from_event_min" in headers:
        row["months_from_event_min"] = parse_number(src.get("months_from_event_min"))
        row["months_from_event_max"] = parse_number(src.get("months_from_event_max"))
        row["months_from_event_basis"] = src.get("months_from_event_basis")
        row["percentage_applies_per"] = src.get("percentage_applies_per") or "once"
        row["source_url"] = src.get("source_url")
    return row


def write_bank_charges(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    # Clear existing content
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for col, h in enumerate(headers, start=1):
        ws.cell(1, col, h)
    for r_i, row in enumerate(rows, start=2):
        for c_i, h in enumerate(headers, start=1):
            ws.cell(r_i, c_i, row.get(h))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--structured", type=Path, default=DEFAULT_STRUCTURED)
    ap.add_argument("--compare", type=Path, default=DEFAULT_COMPARE)
    ap.add_argument("--write-lock-snippet", type=Path, default=None)
    args = ap.parse_args()

    if not args.structured.exists():
        print(f"Missing structured file: {args.structured}", file=sys.stderr)
        return 1
    if not args.compare.exists():
        print(f"Missing compare file: {args.compare}", file=sys.stderr)
        return 1

    structured = read_structured(args.structured)
    # drop not_applicable if any
    structured = [r for r in structured if r.get("charge_status") in ("charged", "nil")]

    wb = load_workbook(args.compare)
    if set(wb.sheetnames) != {"Offers", "Bank_charges", "Government_charges"}:
        print(f"Unexpected sheets: {wb.sheetnames}", file=sys.stderr)
        return 1

    offers_n = sum(
        1
        for row in wb["Offers"].iter_rows(min_row=2, values_only=True)
        if any(c is not None and c != "" for c in row)
    )
    govt_n = sum(
        1
        for row in wb["Government_charges"].iter_rows(min_row=2, values_only=True)
        if any(c is not None and c != "" for c in row)
    )

    ws = wb["Bank_charges"]
    headers, rows = sheet_to_dicts(ws)

    for col in NEW_COLS:
        if col not in headers:
            headers.append(col)
            for r in rows:
                r[col] = None

    before_total = len(rows)
    kept = [r for r in rows if not is_fixed_prepay(r)]
    removed = before_total - len(kept)
    non_fixed_prepay_kept = len(kept)

    next_id = max_pre_id(rows) + 1
    new_rows: list[dict[str, Any]] = []
    for src in structured:
        new_rows.append(structured_to_bank_row(src, headers, next_id))
        next_id += 1

    final_rows = kept + new_rows

    # Safety asserts
    if len([r for r in final_rows if not is_fixed_prepay(r)]) != non_fixed_prepay_kept:
        print("FAIL: non-Fixed-prepay row count changed", file=sys.stderr)
        return 1

    write_bank_charges(ws, headers, final_rows)

    # Re-check offers/govt untouched (same object)
    offers_n2 = sum(
        1
        for row in wb["Offers"].iter_rows(min_row=2, values_only=True)
        if any(c is not None and c != "" for c in row)
    )
    govt_n2 = sum(
        1
        for row in wb["Government_charges"].iter_rows(min_row=2, values_only=True)
        if any(c is not None and c != "" for c in row)
    )
    if offers_n != offers_n2 or govt_n != govt_n2:
        print("FAIL: Offers or Government_charges row count changed", file=sys.stderr)
        return 1

    tmp = args.compare.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    wb.close()
    tmp.replace(args.compare)

    total = len(final_rows)
    print(f"COMPARE updated: {args.compare}")
    print(f"removed_fixed_prepay={removed}")
    print(f"inserted_fixed_prepay={len(new_rows)}")
    print(f"kept_other_bank_charges={non_fixed_prepay_kept}")
    print(f"offers_unchanged={offers_n}")
    print(f"government_unchanged={govt_n}")
    print(f"bank_charges_total={total}")

    if args.write_lock_snippet:
        args.write_lock_snippet.write_text(str(total) + "\n", encoding="utf-8")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

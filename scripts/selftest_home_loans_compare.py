#!/usr/bin/env python3
"""
Self-only query tests against data/HOME_LOANS_COMPARE_v1.xlsx.
Simulates product-style lookups (match Offers / filter Bank_charges / Govt).
Not shipped to customers — run locally after convert.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "HOME_LOANS_COMPARE_v1.xlsx"


def load_sheet(wb, name: str):
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    return headers, rows


def near(a, b, tol=1e-12) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return a == b


class T:
    def __init__(self) -> None:
        self.passed = 0
        self.failed: list[str] = []

    def check(self, name: str, ok: bool, detail: str = "") -> None:
        if ok:
            self.passed += 1
            print(f"  PASS  {name}")
        else:
            self.failed.append(name)
            print(f"  FAIL  {name}  {detail}")


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX}", file=sys.stderr)
        return 1

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    t = T()

    print("=== Sheet shape ===")
    t.check("exactly_3_sheets", wb.sheetnames == ["Offers", "Bank_charges", "Government_charges"], str(wb.sheetnames))

    oh, offers = load_sheet(wb, "Offers")
    bh, banks = load_sheet(wb, "Bank_charges")
    gh, govt = load_sheet(wb, "Government_charges")

    t.check("offers_806", len(offers) == 806, str(len(offers)))
    t.check("bank_1398", len(banks) == 1398, str(len(banks)))
    t.check("govt_18", len(govt) == 18, str(len(govt)))
    t.check("offers_no_fee_cols", not any("processing_fee" in h or h.startswith("overdue_") or "prepayment" in h for h in oh), "")
    t.check("offers_has_roi", "roi" in oh and "bank_key" in oh, "")
    t.check("offers_has_bank_type", "bank_type" in oh and oh.index("bank_type") == oh.index("bank_name") + 1, "")

    print("\n=== Query: Axis Floating → Fixed switch fee (user-style) ===")
    axis_switch = [
        r
        for r in banks
        if r.get("bank_name") == "Axis Bank"
        and r.get("charge_name") == "Interest Rate Type Switch Fees"
        and r.get("interest_rate_type_switch_from") == "Floating"
        and r.get("interest_rate_type_switch_to") == "Fixed"
    ]
    t.check("axis_float_to_fixed_count", len(axis_switch) == 2, str(len(axis_switch)))
    for r in axis_switch:
        fac = r.get("facility_type")
        t.check(
            f"axis_float_to_fixed_{fac}",
            near(r.get("percentage"), 0.01) and r.get("charge_min") == 10000 and r.get("when_it_matters") == "After offer",
            f"pct={r.get('percentage')} min={r.get('charge_min')} when={r.get('when_it_matters')}",
        )

    print("\n=== Query: Axis Fixed → Floating (contrast) ===")
    axis_rev = [
        r
        for r in banks
        if r.get("bank_name") == "Axis Bank"
        and r.get("charge_name") == "Interest Rate Type Switch Fees"
        and r.get("interest_rate_type_switch_from") == "Fixed"
        and r.get("interest_rate_type_switch_to") == "Floating"
    ]
    t.check("axis_fixed_to_float_count", len(axis_rev) == 2, str(len(axis_rev)))
    t.check(
        "axis_fixed_to_float_2pct",
        all(near(r.get("percentage"), 0.02) for r in axis_rev),
        str([r.get("percentage") for r in axis_rev]),
    )

    print("\n=== Query: Offers match → join Bank_charges (Axis Salaried Floating CIBIL 820) ===")
    matched = []
    for o in offers:
        if "axis" not in str(o.get("bank_key") or ""):
            continue
        occ = o.get("occupation")
        if occ not in (None, "", "Any", "Salaried") or o.get("rate_type") != "Floating":
            continue
        st = o.get("cibil_score_status")
        if st == "Not_Used":
            matched.append(o)
        elif st == "Scored":
            try:
                if float(o["cibil_band_score_min"]) <= 820 <= float(o["cibil_band_score_max"]):
                    matched.append(o)
            except (TypeError, ValueError, KeyError):
                pass
    t.check("axis_offers_match_ge1", len(matched) >= 1, str(len(matched)))
    if matched:
        bk = matched[0]["bank_key"]
        roi = matched[0].get("roi")
        t.check("axis_match_has_roi", roi is not None and float(roi) > 0, str(roi))
        joined = [r for r in banks if r.get("bank_key") == bk]
        t.check("axis_join_bank_charges", len(joined) > 0, str(len(joined)))
        proc = [r for r in joined if r.get("charge_name") == "Processing fee"]
        t.check("axis_has_processing_fee_rows", len(proc) >= 1, str(len(proc)))
        before = [r for r in joined if r.get("when_it_matters") == "Before offer"]
        t.check("axis_has_before_offer_fees", len(before) >= 1, str(len(before)))

    print("\n=== Spot: BoB ECS Metro vs Rural ===")
    bob51 = next((r for r in banks if r.get("source_ref") == "Other charges!51"), None)
    bob56 = next((r for r in banks if r.get("source_ref") == "Other charges!56"), None)
    t.check("bob_ecs_metro_125", bob51 is not None and bob51.get("fixed_amount") == 125, str(bob51 and bob51.get("fixed_amount")))
    t.check("bob_ecs_rural_450", bob56 is not None and bob56.get("fixed_amount") == 450, str(bob56 and bob56.get("fixed_amount")))

    print("\n=== Spot: BoI CIC 50 vs 200 ===")
    b62 = next((r for r in banks if r.get("source_ref") == "Other charges!62"), None)
    b63 = next((r for r in banks if r.get("source_ref") == "Other charges!63"), None)
    t.check("boi_cic_50", b62 is not None and b62.get("fixed_amount") == 50, str(b62 and b62.get("fixed_amount")))
    t.check("boi_cic_200", b63 is not None and b63.get("fixed_amount") == 200, str(b63 and b63.get("fixed_amount")))

    print("\n=== Spot: Axis Term Loan repricing slabs ===")
    slabs = [next((r for r in banks if r.get("source_ref") == f"Other charges!{n}"), None) for n in range(16, 20)]
    t.check(
        "axis_repricing_1000_2000_3000_5000",
        all(s is not None for s in slabs) and [s.get("fixed_amount") for s in slabs] == [1000, 2000, 3000, 5000],
        str([s.get("fixed_amount") if s else None for s in slabs]),
    )
    if slabs[0]:
        gid = slabs[0]["charge_group_id"]
        group = [r for r in banks if r.get("charge_group_id") == gid]
        t.check("axis_repricing_group_4", len(group) == 4, str(len(group)))
        # ordered by slab_from
        froms = [r.get("slab_from") for r in group]
        t.check("axis_repricing_slab_ordered", froms == sorted(froms, key=lambda x: (x is None, x)), str(froms))

    print("\n=== Spot: PNB processing 0 vs 0.35% ===")
    pnb_proc = [
        r
        for r in banks
        if r.get("origin") == "Offers.processing" and "punjab national" in str(r.get("bank_key") or "")
    ]
    pcts = {r.get("percentage") for r in pnb_proc}
    t.check("pnb_proc_has_0", 0 in pcts or 0.0 in pcts, str(pcts))
    t.check("pnb_proc_has_0_0035", any(near(p, 0.0035) for p in pcts if p is not None), str(pcts))

    print("\n=== Spot: DCB overdue Slab_Table ladder ===")
    dcb = [r for r in banks if r.get("origin") == "Slab_Table"]
    t.check("dcb_slab_76", len(dcb) == 76, str(len(dcb)))
    dcb_sorted = sorted(dcb, key=lambda r: float(r.get("slab_from") or 0))
    first = dcb_sorted[0]
    t.check(
        "dcb_first_0_3000_150",
        first.get("slab_from") == 0 and first.get("slab_to") == 3000 and first.get("fixed_amount") == 150,
        str((first.get("slab_from"), first.get("slab_to"), first.get("fixed_amount"))),
    )
    t.check("dcb_one_charge_group", len({r.get("charge_group_id") for r in dcb}) == 1, "")

    print("\n=== Spot: Canara overdue bands exist ===")
    canara_od = [
        r
        for r in banks
        if r.get("origin") == "Offers.overdue" and "canara" in str(r.get("bank_key") or "")
    ]
    t.check("canara_overdue_ge1", len(canara_od) >= 1, str(len(canara_od)))

    print("\n=== Spot: Floating prepay mostly not charged ===")
    float_prepay = [
        r
        for r in banks
        if r.get("charge_name") == "Prepayment charges" and r.get("rate_type") == "Floating"
    ]
    if float_prepay:
        zeroish = sum(1 for r in float_prepay if r.get("fixed_amount") == 0)
        t.check("float_prepay_mostly_zero", zeroish / len(float_prepay) >= 0.5, f"{zeroish}/{len(float_prepay)}")
    else:
        t.check("float_prepay_mostly_zero", False, "no floating prepay rows")

    print("\n=== Spot: Government unit + CERSAI ===")
    t.check(
        "govt_no_loan_agreement_stamp",
        not any(r.get("charge_name") == "Loan Agreement Stamp Duty" for r in govt),
        "",
    )
    modt = [r for r in govt if r.get("charge_name") == "MODT Stamp Duty"]
    t.check("govt_modt_exists", len(modt) >= 1, str(len(modt)))
    t.check(
        "govt_modt_first_pct_0_001",
        modt and near(modt[0].get("percentage"), 0.001),
        str(modt[0].get("percentage") if modt else None),
    )
    cersai = [r for r in govt if r.get("charge_name") and "CERSAI" in str(r.get("charge_name"))]
    t.check("cersai_flat_50", any(r.get("flat_amount_inr") == 50 for r in cersai), str([(r.get("charge_name"), r.get("flat_amount_inr")) for r in cersai]))

    print("\n=== Spot: purpose Home Loan gone; NEEDS_REVIEW none ===")
    t.check("no_purpose_home_loan", not any(r.get("purpose") == "Home Loan" for r in banks), "")
    nr = {r.get("charge_name") for r in banks if r.get("when_it_matters") == "NEEDS_REVIEW"}
    t.check("needs_review_none", len(nr) == 0, str(nr))

    print("\n=== EMI ingredients check (match has amount/tenure/roi) ===")
    if matched:
        o = matched[0]
        # Can compute EMI if we have roi and customer amount/tenure fall in bands (or bands not used)
        can = o.get("roi") is not None
        t.check("emi_ingredients_roi_present", can, str(o.get("roi")))
        t.check(
            "emi_ingredients_tenure_cols",
            "tenure_months_min" in o and "tenure_months_max" in o,
            "",
        )

    wb.close()

    print("\n" + "=" * 50)
    print(f"Passed: {t.passed}   Failed: {len(t.failed)}")
    if t.failed:
        print("Failed tests:", ", ".join(t.failed))
        return 1
    print("All self-tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

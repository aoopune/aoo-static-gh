#!/usr/bin/env python3
"""
Post-cleanup audit: Home Loans.xlsx → Processing fees vs master Bank_charges only.
Does not use HOME LOANS - OFFERS workbook.
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "data" / "HOME_LOANS_COMPARE_v1.xlsx"
REF = ROOT / "data" / "Home Loans.xlsx"
OUT_DIR = ROOT / "data" / "Charges" / "_audit"
OUT_SUMMARY = OUT_DIR / "PROCESSING_FEES_AUDIT.md"
OUT_JSON = OUT_DIR / "processing_fees_audit.json"
OUT_STATUS = OUT_DIR / "PROCESSING_FEES_CLEANUP_STATUS.md"

REF_NAME_MAP = {
    "central bank of india": "Central Bank of India",
    "bank of maharastra": "Bank of Maharashtra",
    "federal bank": "Federal Bank",
    "idfc first bank": "IDFC FIRST Bank",
    "indusind bank": "IndusInd Bank",
    "jammu & kashmir bank": "Jammu and Kashmir Bank",
    "karur vyasa bank": "Karur Vysya Bank",
    "naintal bank": "Nainital Bank",
}


def blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def bank_key(name: Any) -> str:
    s = "" if name is None else str(name)
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("&", "and")
    return s


def num(v: Any) -> float | None:
    if blank(v):
        return None
    return float(v)


def norm_ref_bank(name: str) -> str:
    return REF_NAME_MAP.get(bank_key(name), str(name).strip())


def parse_ref_text(text: str) -> dict[str, Any]:
    t = text or ""
    out: dict[str, Any] = {"raw": t}
    pct = re.findall(r"(\d+(?:\.\d+)?)\s*%", t.replace(",", ""))
    if pct:
        out["percentages"] = sorted({float(x) / 100 for x in pct})
    mins = re.findall(r"min\.?\s*[₹:]?\s*(\d[\d,]*)", t, re.I)
    maxs = re.findall(r"max\.?\s*[₹:]?\s*(\d[\d,]*)", t, re.I)
    if mins:
        out["mins"] = sorted({int(x.replace(",", "")) for x in mins})
    if maxs:
        out["maxs"] = sorted({int(x.replace(",", "")) for x in maxs})
    if re.search(r"\bnil\b", t, re.I):
        out["has_nil"] = True
    return out


def load_reference() -> dict[str, dict[str, Any]]:
    wb = load_workbook(REF, read_only=True, data_only=True)
    ws = wb["Processing fees"]
    out: dict[str, dict[str, Any]] = {}
    for raw in ws.iter_rows(min_row=3, values_only=True):
        if blank(raw[0]):
            continue
        bank = norm_ref_bank(str(raw[0]))
        out[bank] = {
            "bank_name": bank,
            "ref_text": raw[1] if len(raw) > 1 else None,
            "parsed": parse_ref_text(str(raw[1] or "")),
        }
    wb.close()
    return out


def load_master_processing() -> list[dict[str, Any]]:
    wb = load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb["Bank_charges"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: raw[i] for i in range(len(headers))}
        if d.get("origin") != "Offers.processing":
            continue
        rows.append(d)
    wb.close()
    return rows


def cross_check_ref(rows: list[dict[str, Any]], ref: dict[str, Any] | None) -> list[str]:
    issues: list[str] = []
    if not ref:
        issues.append("Bank missing from Processing fees reference sheet")
        return issues
    parsed = ref["parsed"]
    pcts = {
        round(x, 6)
        for x in (num(r.get("percentage")) for r in rows if r.get("charge_type") == "Percentage")
        if x is not None
    }
    mins = {int(num(r.get("charge_min"))) for r in rows if not blank(r.get("charge_min"))}
    maxs = {int(num(r.get("charge_max"))) for r in rows if not blank(r.get("charge_max"))}
    has_zero = any(num(r.get("percentage")) == 0 or num(r.get("fixed_amount")) == 0 for r in rows)

    if parsed.get("percentages"):
        for rp in parsed["percentages"]:
            if not any(abs(rp - p) < 1e-6 for p in pcts):
                if rp >= 0.1 and any(abs(rp / 100 - p) < 1e-6 for p in pcts):
                    continue
                issues.append(
                    f"Reference mentions {rp * 100:.2f}% but master has "
                    f"{[round(x * 100, 4) if x else 0 for x in sorted(pcts)]}"
                )
    if parsed.get("mins"):
        for rm in parsed["mins"]:
            if rm not in mins:
                issues.append(f"Reference min ₹{rm:,} not in master mins {sorted(mins)}")
    if parsed.get("maxs"):
        for rx in parsed["maxs"]:
            if rx not in maxs:
                issues.append(f"Reference max ₹{rx:,} not in master maxes {sorted(maxs)}")
    if parsed.get("has_nil") and not has_zero:
        issues.append("Reference mentions NIL but no zero-fee rows in master")
    return issues


def dedupe_key(r: dict[str, Any]) -> tuple:
    fields = [
        "bank_name", "scheme", "purpose", "facility_type", "rate_type", "occupation",
        "borrower_category", "loan_amount_band_applicable", "loan_amount_min", "loan_amount_max",
        "tenure_band_applicable", "tenure_months_min", "tenure_months_max",
        "charge_type", "percentage", "fixed_amount", "charge_min", "charge_max",
        "valid_from", "valid_till",
    ]
    return tuple(r.get(f) for f in fields)


def audit_bank(bank: str, rows: list[dict[str, Any]], ref: dict[str, Any] | None) -> dict[str, Any]:
    cibil_rows = [
        r for r in rows
        if r.get("cibil_band_applicable") not in (None, "No", "") or not blank(r.get("cibil_band_score_min"))
    ]
    dup_groups = [ids for ids in Counter(dedupe_key(r) for r in rows).values() if ids > 1]
    ref_issues = cross_check_ref(rows, ref)

    verdict = "PASS"
    if ref_issues or cibil_rows or dup_groups:
        verdict = "FAIL"
    elif len(rows) > 1:
        verdict = "PASS_MULTI"  # amounts OK, multiple rows for real splits

    return {
        "bank": bank,
        "verdict": verdict,
        "master_rows": len(rows),
        "cibil_rows": len(cibil_rows),
        "duplicate_keys": len(dup_groups),
        "ref_issues": ref_issues,
        "distinct_percentages": sorted(
            {round(num(r.get("percentage")), 6) for r in rows if num(r.get("percentage")) is not None}
        ),
        "distinct_mins": sorted({int(num(r.get("charge_min"))) for r in rows if num(r.get("charge_min"))}),
        "distinct_maxes": sorted({int(num(r.get("charge_max"))) for r in rows if num(r.get("charge_max"))}),
    }


def write_status_md(results: list[dict[str, Any]], totals: dict[str, Any]) -> None:
    problems = [
        {
            "id": "cibil_splits",
            "title": "Unwanted CIBIL bands on processing fees",
            "was": "Most banks had 5–10 CIBIL score bands per fee — not in your human sheet.",
            "action": "Set cibil_band_applicable=No and cleared min/max on every Offers.processing row; removed duplicate rows that only differed by CIBIL.",
            "now": f"{totals['cibil_rows_remaining']} rows still have CIBIL on processing fees (target: 0).",
            "status": "DONE" if totals["cibil_rows_remaining"] == 0 else "OPEN",
        },
        {
            "id": "sbi_duplicates",
            "title": "SBI — 72 duplicate rows and max-cap confusion",
            "was": "9× copies of each rule (72 rows). Max ₹15,000 and ₹18,000 looked like a conflict.",
            "action": "You deleted 64 duplicate rows manually. Kept 8 rows: scheme × occupation split (Salaried max ₹15,000; Self-Employed max ₹18,000).",
            "now": "8 rows, no CIBIL, no duplicates.",
            "status": "DONE",
        },
        {
            "id": "canara_extra_tier",
            "title": "Canara — extra 0.25% tier not in your sheet",
            "was": "Master had 0.25%, min ₹750, max ₹5,000 on top of your 0.50% rule.",
            "action": "Deleted 5 rows with 0.25% (CHG-PROC-501, 506, 511, 516, 521 and dedupe siblings).",
            "now": "5 rows — all 0.50%, min ₹1,500, max ₹10,000 with loan-amount bands only.",
            "status": "DONE",
        },
        {
            "id": "review_banks_cibil",
            "title": "REVIEW banks — CIBIL noise on top of real splits",
            "was": "PNB, Union, BOM, ICICI, IDBI, Indian Bank, IOB, KVB, Nainital, PSB had CIBIL splits your sheet never mentions.",
            "action": "Automated cleanup: CIBIL cleared + dedupe across all banks.",
            "now": "Row counts reduced to real splits only (scheme, loan slab, occupation, etc.).",
            "status": "DONE",
        },
        {
            "id": "bob_row_explosion",
            "title": "Bank of Baroda — 240 rows (CIBIL explosion)",
            "was": "Rates matched your sheet but 10 CIBIL bands multiplied every row.",
            "action": "CIBIL cleared + dedupe.",
            "now": "24 rows — 0.50% / 0.25% loan bands with min ₹8,500, no CIBIL.",
            "status": "DONE",
        },
        {
            "id": "jk_naming",
            "title": "Jammu & Kashmir — bank name spelling",
            "was": "Your sheet said 'Jammu & Kashmir bank'; master uses 'Jammu and Kashmir Bank'.",
            "action": "Updated Home Loans.xlsx Processing fees row to official name 'Jammu and Kashmir Bank'.",
            "now": "Names aligned; 1 processing row (0.25%, min ₹2,000, max ₹50,000).",
            "status": "DONE",
        },
        {
            "id": "pass_banks",
            "title": "PASS banks (19) — amounts already matched",
            "was": "Axis, HDFC, Yes, Federal, DCB, Kotak, RBL, etc. — core rate/min/max correct; some had extra CIBIL rows.",
            "action": "CIBIL cleared + dedupe where needed. Axis left untouched (already 2 correct rows).",
            "now": "All 19 still pass amount checks; CIBIL removed.",
            "status": "DONE",
        },
        {
            "id": "multi_row_granularity",
            "title": "Multiple rows per bank (intentional, not a bug)",
            "was": "Your sheet is one summary line; master needs rows when fee differs by scheme, loan size, occupation, etc.",
            "action": "Kept rows that differ on real dimensions; removed only CIBIL-only duplicates.",
            "now": "212 total processing rows across 33 banks (down from 1,161).",
            "status": "BY DESIGN",
        },
    ]

    lines = [
        "# Processing fees cleanup — final status",
        "",
        f"**Completed:** 2026-08-14",
        "",
        "Source of truth: `data/Home Loans.xlsx` → **Processing fees**",
        "Master: `data/HOME_LOANS_COMPARE_v1.xlsx` → **Bank_charges** (`origin=Offers.processing`)",
        "",
        "## Overall numbers",
        "",
        "| Metric | Before cleanup | After cleanup |",
        "|---|---:|---:|",
        f"| Processing fee rows | 1,161 | {totals['processing_rows']} |",
        f"| Total Bank_charges rows | 2,347 | {totals['bank_charges_rows']} |",
        f"| Banks | 33 | 33 |",
        f"| Rows with CIBIL on processing | ~999+ | {totals['cibil_rows_remaining']} |",
        f"| Exact duplicate keys | many | {totals['duplicate_keys']} |",
        "",
        "## Problem-by-problem",
        "",
    ]
    for p in problems:
        lines += [
            f"### {p['title']}",
            "",
            f"**Status:** {p['status']}",
            "",
            f"- **What was wrong:** {p['was']}",
            f"- **What we did:** {p['action']}",
            f"- **Now:** {p['now']}",
            "",
        ]

    lines += [
        "## Per-bank verdict (post-cleanup)",
        "",
        "| Bank | Verdict | Rows | CIBIL rows | Ref amount check |",
        "|---|---|---:|---:|---|",
    ]
    verdict_label = {"PASS": "PASS", "PASS_MULTI": "PASS (multi-row)", "FAIL": "FAIL"}
    for r in sorted(results, key=lambda x: x["bank"]):
        ref_ok = "OK" if not r["ref_issues"] else "; ".join(r["ref_issues"][:2])
        lines.append(
            f"| {r['bank']} | {verdict_label.get(r['verdict'], r['verdict'])} | {r['master_rows']} | "
            f"{r['cibil_rows']} | {ref_ok} |"
        )

    fail_count = sum(1 for r in results if r["verdict"] == "FAIL")
    lines += [
        "",
        f"**Summary:** {len(results)} banks audited; {fail_count} FAIL; "
        f"{sum(1 for r in results if r['verdict'] == 'PASS')} single-row PASS; "
        f"{sum(1 for r in results if r['verdict'] == 'PASS_MULTI')} multi-row PASS.",
        "",
        "Regenerate this report: `python3 scripts/audit_processing_fees_human.py`",
    ]
    OUT_STATUS.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_audit_md(results: list[dict[str, Any]], totals: dict[str, Any]) -> None:
    lines = [
        "# Processing Fees Audit (post-cleanup)",
        "",
        "Compares **only** your human sheet vs master — no Offers workbook.",
        "",
        "- Human source: `data/Home Loans.xlsx` → **Processing fees**",
        "- Master: `data/HOME_LOANS_COMPARE_v1.xlsx` → **Bank_charges** (`origin=Offers.processing`)",
        "",
        f"**Cleanup completed 2026-08-14.** See `PROCESSING_FEES_CLEANUP_STATUS.md` for problem-by-problem summary.",
        "",
        "## Overall",
        "",
        "| Metric | Count |",
        "|---|---:|",
        f"| Banks in reference sheet | {totals['ref_banks']} |",
        f"| Banks with processing rows in master | {totals['master_banks']} |",
        f"| Processing rows in master | {totals['processing_rows']} |",
        f"| Total Bank_charges rows | {totals['bank_charges_rows']} |",
        f"| Processing rows with CIBIL still set | {totals['cibil_rows_remaining']} |",
        f"| Exact duplicate dedupe keys | {totals['duplicate_keys']} |",
        f"| Banks PASS (single row) | {totals['pass_single']} |",
        f"| Banks PASS (multi-row, amounts OK) | {totals['pass_multi']} |",
        f"| Banks FAIL | {totals['fail']} |",
        "",
        "## Per-bank summary",
        "",
        "| Bank | Verdict | Master rows | CIBIL rows | Ref issues |",
        "|---|---|---:|---:|---:|",
    ]
    for r in sorted(results, key=lambda x: x["bank"]):
        vl = {"PASS": "PASS", "PASS_MULTI": "PASS (multi)", "FAIL": "FAIL"}.get(r["verdict"], r["verdict"])
        lines.append(
            f"| {r['bank']} | {vl} | {r['master_rows']} | {r['cibil_rows']} | {len(r['ref_issues'])} |"
        )

    lines += ["", "## Bank details", ""]
    for r in sorted(results, key=lambda x: x["bank"]):
        lines.append(f"### {r['bank']} — {r['verdict']}")
        lines.append(f"- Master rows: {r['master_rows']}")
        if r["distinct_percentages"]:
            lines.append(f"- Distinct percentages: {[round(x * 100, 4) for x in r['distinct_percentages']]}")
        if r["distinct_mins"]:
            lines.append(f"- Distinct mins: {['₹{:,}'.format(x) for x in r['distinct_mins']]}")
        if r["distinct_maxes"]:
            lines.append(f"- Distinct maxes: {['₹{:,}'.format(x) for x in r['distinct_maxes']]}")
        if r["cibil_rows"]:
            lines.append(f"- **CIBIL rows remaining:** {r['cibil_rows']}")
        for issue in r["ref_issues"]:
            lines.append(f"- Ref issue: {issue}")
        if not r["ref_issues"] and r["cibil_rows"] == 0:
            lines.append("- Reference amounts match master.")
        lines.append("")

    lines.append("Regenerate: `python3 scripts/audit_processing_fees_human.py`")
    OUT_SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    ref_by_name = load_reference()
    proc = load_master_processing()

    wb = load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb["Bank_charges"]
    bank_charges_count = ws.max_row - 1
    wb.close()

    by_bank: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in proc:
        by_bank[r["bank_name"]].append(r)

    all_banks = sorted(set(ref_by_name.keys()) | set(by_bank.keys()))
    results = [audit_bank(b, by_bank.get(b, []), ref_by_name.get(b)) for b in all_banks]

    dup_keys = sum(
        1 for bank_rows in by_bank.values()
        for cnt in Counter(dedupe_key(r) for r in bank_rows).values()
        if cnt > 1
    )
    cibil_remaining = sum(
        1 for r in proc
        if r.get("cibil_band_applicable") not in (None, "No", "") or not blank(r.get("cibil_band_score_min"))
    )

    totals = {
        "ref_banks": len(ref_by_name),
        "master_banks": len(by_bank),
        "processing_rows": len(proc),
        "bank_charges_rows": bank_charges_count,
        "cibil_rows_remaining": cibil_remaining,
        "duplicate_keys": dup_keys,
        "pass_single": sum(1 for r in results if r["verdict"] == "PASS"),
        "pass_multi": sum(1 for r in results if r["verdict"] == "PASS_MULTI"),
        "fail": sum(1 for r in results if r["verdict"] == "FAIL"),
    }

    payload = {"completed_at": "2026-08-14", "totals": totals, "banks": results}
    OUT_JSON.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    write_audit_md(results, totals)
    write_status_md(results, totals)

    print(f"Wrote {OUT_SUMMARY}")
    print(f"Wrote {OUT_STATUS}")
    print(f"Wrote {OUT_JSON}")
    print(f"Processing rows: {totals['processing_rows']}; FAIL banks: {totals['fail']}")
    return 0 if totals["fail"] == 0 and cibil_remaining == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())

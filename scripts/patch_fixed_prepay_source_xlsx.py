#!/usr/bin/env python3
"""Apply fixed prepayment source fixes to Home Loans.xlsx (single source of truth)."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data" / "Home Loans.xlsx"
SHEET = "Pre-payment charges - takeoverf"

# row -> (bank_name, tl_self, tl_take, od_self, od_take, source_url)
PATCHES: dict[int, tuple[str, str, str, str, str, str | None]] = {
    9: (
        "Indian Bank",
        "NIL",
        "2% on the amount being paid",
        "NIL",
        "2% on the amount being paid",
        "https://indianbank.bank.in/documents/d/guest/annexure_circular-web-wef-01-01-26-1-1",
    ),
    10: (
        "Canara Bank",
        "NIL",
        "2% on the amount being paid",
        "NIL",
        "2% on the amount being paid",
        "https://www.canarabank.bank.in/documents/d/guest/policy-on-penal-charges-101125",
    ),
    13: ("State Bank of India", "NA", "NA", "NA", "NA", None),
    14: (
        "Bank of Maharashtra",
        "NIL",
        "NIL",
        "NA",
        "NA",
        "https://bankofmaharashtra.bank.in/service-charges",
    ),
    18: (
        "City Union Bank",
        "NIL",
        "2% on the amount being paid",
        "NA",
        "NA",
        "https://cityunionbank.bank.in/service-charges",
    ),
    29: (
        "Kotak Mahindra Bank",
        "NIL",
        "1% on outstanding amount of each year of residual period to original maturity",
        "NIL",
        "1% on outstanding amount of each year of residual period to original maturity",
        "https://www.kotak.bank.in/en/personal-banking/loans/home-loan/fees-and-charges.html",
    ),
    35: ("Tamilnad Mercantile Bank", "NA", "NA", "NA", "NA", None),
}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = ap.parse_args()
    if not args.xlsx.exists():
        print(f"Missing xlsx: {args.xlsx}", file=sys.stderr)
        return 1

    wb = load_workbook(args.xlsx)
    if SHEET not in wb.sheetnames:
        print(f"Missing sheet {SHEET!r}", file=sys.stderr)
        return 1
    ws = wb[SHEET]

    for row_num, (bank, tl_self, tl_take, od_self, od_take, url) in PATCHES.items():
        ws.cell(row_num, 1, bank)
        ws.cell(row_num, 2, tl_self)
        ws.cell(row_num, 3, tl_take)
        ws.cell(row_num, 4, od_self)
        ws.cell(row_num, 5, od_take)
        ws.cell(row_num, 6, url if url else None)

    wb.save(args.xlsx)
    wb.close()
    print(f"Patched {len(PATCHES)} bank rows in {args.xlsx}!{SHEET}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

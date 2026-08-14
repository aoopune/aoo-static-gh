#!/usr/bin/env python3
"""
Parse fixed-rate prepayment source (Home Loans.xlsx sheet or legacy CSV)
→ data/fixed-prepay-structured.csv. Atomic rows only. No COMPARE writes.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data" / "Home Loans.xlsx"
DEFAULT_SHEET = "Pre-payment charges - takeoverf"
DEFAULT_INPUT = ROOT / "data" / "Home Loans - Pre-payment charges - takeover_fixed rrate (1).csv"
DEFAULT_OUTPUT = ROOT / "data" / "fixed-prepay-structured.csv"

BANK_ALIASES = {
    "Central bank of India": "Central Bank of India",
    "Bank of Maharastra": "Bank of Maharashtra",
    "Federal bank": "Federal Bank",
    "IDFC first Bank": "IDFC FIRST Bank",
    "Indusind Bank": "IndusInd Bank",
    "Karur Vyasa Bank": "Karur Vysya Bank",
    "Jammu & Kashmir bank": "Jammu and Kashmir Bank",
}

MASTER_HEADERS = [
    "bank_name",
    "bank_key",
    "facility_type",
    "charge_name",
    "rate_type",
    "charge_status",
    "percentage",
    "percentage_base_value",
    "fixed_amount",
    "gst_applicable",
    "has_slab_wise_charges",
    "slab_from",
    "slab_to",
    "slab_basis",
    "loan_amount_band_applicable",
    "loan_amount_min",
    "loan_amount_max",
    "months_from_event_min",
    "months_from_event_max",
    "months_from_event_basis",
    "percentage_applies_per",
    "source_url",
    "note_1",
    "csv_cell_ref",
]


def bank_key(name: str) -> str:
    s = (name or "").lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s.replace("&", "and")


def norm_bank(raw: str) -> str:
    s = re.sub(r"\s+", " ", (raw or "").strip())
    return BANK_ALIASES.get(s, s)


def normalize_source_url(raw: str | None) -> str | None:
    s = (raw or "").strip().replace("\n", " | ")
    if not s:
        return None
    if s.lower().startswith(("http://", "https://")):
        return s
    return None


def cell_kind(v: str) -> str:
    t = (v or "").strip()
    if not t:
        return "blank"
    u = t.upper().replace(" ", "")
    if u in ("NIL", "NIL."):
        return "nil"
    if u in ("NA", "N/A", "N.A."):
        return "na"
    return "rule"


def base_from_text(text: str) -> str:
    t = text.lower()
    if "prepaid" in t and "amount being paid" not in t:
        return "Prepaid_Amount"
    if "sanctioned" in t:
        return "Sanctioned_Limit"
    if "highest outstanding" in t or "90 days" in t or "90-day" in t:
        return "Highest_Outstanding_90_Days"
    if "o/s" in t or "outstanding" in t or "ouststanding" in t:
        return "Outstanding_Amount"
    if "amount being paid" in t or "amount being prepaid" in t:
        return "Amount_Being_Paid"
    return "Amount_Being_Paid"


def simple_pct(text: str) -> float | None:
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*%", text)
    return float(m.group(1)) / 100.0 if m else None


def empty_row(**kwargs: Any) -> dict[str, Any]:
    row = {h: None for h in MASTER_HEADERS}
    row["rate_type"] = "Fixed"
    row["percentage_applies_per"] = "once"
    row["has_slab_wise_charges"] = "No"
    row.update(kwargs)
    return row


def emit_nil(
    rows: list[dict[str, Any]],
    *,
    bank: str,
    facility: str,
    charge_name: str,
    source_url: str,
    csv_cell_ref: str,
) -> None:
    rows.append(
        empty_row(
            bank_name=bank,
            bank_key=bank_key(bank),
            facility_type=facility,
            charge_name=charge_name,
            charge_status="nil",
            fixed_amount=0,
            gst_applicable=None,
            source_url=source_url or None,
            note_1="Prepayment nil (CSV NIL)",
            csv_cell_ref=csv_cell_ref,
        )
    )


def emit_simple_charged(
    rows: list[dict[str, Any]],
    *,
    bank: str,
    facility: str,
    charge_name: str,
    text: str,
    source_url: str,
    csv_cell_ref: str,
    percentage: float | None = None,
    base: str | None = None,
    percentage_applies_per: str = "once",
    months_min: Any = None,
    months_max: Any = None,
    months_basis: str | None = None,
    slab_from: Any = None,
    slab_to: Any = None,
    slab_basis: str | None = None,
    has_slab: str = "No",
    note_1: str | None = None,
) -> None:
    pct = percentage if percentage is not None else simple_pct(text)
    if pct is None:
        raise ValueError(f"Cannot parse percentage from {text!r} ({csv_cell_ref})")
    loan_band = "Yes" if has_slab == "Yes" else None
    rows.append(
        empty_row(
            bank_name=bank,
            bank_key=bank_key(bank),
            facility_type=facility,
            charge_name=charge_name,
            charge_status="charged",
            percentage=pct,
            percentage_base_value=base or base_from_text(text),
            gst_applicable="Yes",
            has_slab_wise_charges=has_slab,
            slab_from=slab_from,
            slab_to=slab_to,
            slab_basis=slab_basis,
            loan_amount_band_applicable=loan_band,
            loan_amount_min=slab_from if loan_band else None,
            loan_amount_max=slab_to if loan_band else None,
            months_from_event_min=months_min,
            months_from_event_max=months_max,
            months_from_event_basis=months_basis,
            percentage_applies_per=percentage_applies_per,
            source_url=source_url or None,
            note_1=note_1,
            csv_cell_ref=csv_cell_ref,
        )
    )


def expand_cell(
    rows: list[dict[str, Any]],
    *,
    bank: str,
    facility: str,
    charge_name: str,
    raw: str,
    source_url: str,
    csv_cell_ref: str,
) -> None:
    kind = cell_kind(raw)
    if kind == "na":
        return
    if kind == "blank":
        # Only Yes Bank OD blanks are mirrored by caller; other blanks = skip
        return
    if kind == "nil":
        emit_nil(
            rows,
            bank=bank,
            facility=facility,
            charge_name=charge_name,
            source_url=source_url,
            csv_cell_ref=csv_cell_ref,
        )
        return

    text = raw.strip()
    low = text.lower()

    # IDBI own-funds time band
    if "within 6 months" in low and "final disbursement" in low:
        emit_simple_charged(
            rows,
            bank=bank,
            facility=facility,
            charge_name=charge_name,
            text=text,
            source_url=source_url,
            csv_cell_ref=csv_cell_ref,
            percentage=0.02,
            base="Amount_Being_Paid",
            months_min=0,
            months_max=6,
            months_basis="final_disbursement",
            note_1="Within 6 months from final disbursement",
        )
        emit_nil(
            rows,
            bank=bank,
            facility=facility,
            charge_name=charge_name,
            source_url=source_url,
            csv_cell_ref=csv_cell_ref + "|after6m",
        )
        # Patch last nil with months band
        rows[-1]["months_from_event_min"] = 6
        rows[-1]["months_from_event_max"] = None
        rows[-1]["months_from_event_basis"] = "final_disbursement"
        rows[-1]["note_1"] = "After 6 months from final disbursement — nil"
        return

    # Bandhan takeover age bands
    if "less than 12 months" in low and "after 12 months" in low:
        emit_simple_charged(
            rows,
            bank=bank,
            facility=facility,
            charge_name=charge_name,
            text=text,
            source_url=source_url,
            csv_cell_ref=csv_cell_ref + "|lt12",
            percentage=0.04,
            base="Outstanding_Amount",
            months_min=0,
            months_max=12,
            months_basis="disbursement",
            note_1="Loan age under 12 months from disbursement",
        )
        emit_simple_charged(
            rows,
            bank=bank,
            facility=facility,
            charge_name=charge_name,
            text=text,
            source_url=source_url,
            csv_cell_ref=csv_cell_ref + "|ge12",
            percentage=0.02,
            base="Outstanding_Amount",
            months_min=12,
            months_max=None,
            months_basis="disbursement",
            note_1="Loan age 12+ months from disbursement",
        )
        return

    # South Indian amount slabs
    if "crore" in low and ("4%" in text or "4 %" in text):
        emit_simple_charged(
            rows,
            bank=bank,
            facility=facility,
            charge_name=charge_name,
            text=text,
            source_url=source_url,
            csv_cell_ref=csv_cell_ref + "|le1cr",
            percentage=0.04,
            base="Amount_Being_Paid",
            has_slab="Yes",
            slab_from=0,
            slab_to=10000000,
            slab_basis="loan_amount",
            note_1="Up to Rs 1 crore loan amount",
        )
        emit_simple_charged(
            rows,
            bank=bank,
            facility=facility,
            charge_name=charge_name,
            text=text,
            source_url=source_url,
            csv_cell_ref=csv_cell_ref + "|gt1cr",
            percentage=0.03,
            base="Amount_Being_Paid",
            has_slab="Yes",
            slab_from=10000000.01,
            slab_to=None,
            slab_basis="loan_amount",
            note_1="Above Rs 1 crore loan amount",
        )
        return

    # Kotak residual years
    if "residual" in low:
        emit_simple_charged(
            rows,
            bank=bank,
            facility=facility,
            charge_name=charge_name,
            text=text,
            source_url=source_url,
            csv_cell_ref=csv_cell_ref,
            percentage=0.01,
            base="Outstanding_Amount",
            percentage_applies_per="residual_year_to_original_maturity",
            note_1="1% of outstanding per residual year to original maturity",
        )
        return

    # Simple percentage rule
    emit_simple_charged(
        rows,
        bank=bank,
        facility=facility,
        charge_name=charge_name,
        text=text,
        source_url=source_url,
        csv_cell_ref=csv_cell_ref,
    )


def _cell_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def structure_bank_rows(
    bank_rows: list[tuple[str, list[str], str | None]],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    banks_seen: list[str] = []

    for bank, cells_raw, source_url in bank_rows:
        banks_seen.append(bank)
        while len(cells_raw) < 4:
            cells_raw.append("")

        cells = [
            ("Term Loan", "Prepayment charges", cells_raw[0], "TL_self"),
            ("Term Loan", "Prepayment charges (takeover)", cells_raw[1], "TL_take"),
            ("Overdraft", "Prepayment charges", cells_raw[2], "OD_self"),
            ("Overdraft", "Prepayment charges (takeover)", cells_raw[3], "OD_take"),
        ]

        if bank == "Yes Bank":
            if cell_kind(cells_raw[2]) == "blank":
                cells[2] = ("Overdraft", "Prepayment charges", cells_raw[0], "OD_self_mirrored_TL")
            if cell_kind(cells_raw[3]) == "blank":
                cells[3] = (
                    "Overdraft",
                    "Prepayment charges (takeover)",
                    cells_raw[1],
                    "OD_take_mirrored_TL",
                )

        for facility, charge_name, raw, ref in cells:
            expand_cell(
                out,
                bank=bank,
                facility=facility,
                charge_name=charge_name,
                raw=raw,
                source_url=source_url or "",
                csv_cell_ref=f"{bank}|{ref}",
            )

    if len(set(banks_seen)) != 33:
        raise SystemExit(
            f"Expected 33 banks, got {len(set(banks_seen))}: {sorted(set(banks_seen))}"
        )
    return out


def load_bank_rows_from_xlsx(path: Path, sheet_name: str) -> list[tuple[str, list[str], str | None]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name!r} in {path}")
    ws = wb[sheet_name]
    bank_rows: list[tuple[str, list[str], str | None]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        bank = norm_bank(_cell_text(row[0]))
        cells = [_cell_text(row[i]) if len(row) > i else "" for i in range(1, 5)]
        url = normalize_source_url(_cell_text(row[5]) if len(row) > 5 else "")
        bank_rows.append((bank, cells, url))
    wb.close()
    return bank_rows


def load_bank_rows_from_csv(path: Path) -> list[tuple[str, list[str], str | None]]:
    raw_rows = list(csv.reader(path.open(encoding="utf-8")))
    bank_rows: list[tuple[str, list[str], str | None]] = []
    for r in raw_rows[3:]:
        if not r or not (r[0] or "").strip():
            continue
        while len(r) < 6:
            r.append("")
        bank = norm_bank(r[0])
        cells = [r[1], r[2], r[3], r[4]]
        url = normalize_source_url(r[5])
        bank_rows.append((bank, cells, url))
    return bank_rows


def structure(input_path: Path, *, from_xlsx: bool = False, sheet_name: str = DEFAULT_SHEET) -> list[dict[str, Any]]:
    if from_xlsx:
        bank_rows = load_bank_rows_from_xlsx(input_path, sheet_name)
    else:
        bank_rows = load_bank_rows_from_csv(input_path)
    return structure_bank_rows(bank_rows)


def export_prepay_sheet_to_csv(
    xlsx_path: Path,
    csv_path: Path,
    *,
    sheet_name: str = DEFAULT_SHEET,
) -> None:
    bank_rows = load_bank_rows_from_xlsx(xlsx_path, sheet_name)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "RBI guidelines",
                "https://rbidocs.rbi.org.in/rdocs/notification/PDFs/170MD.pdf",
                "",
                "",
                "",
                "",
            ]
        )
        w.writerow(["", "Term loan", "", "OD facility'", "", ""])
        w.writerow(
            [
                "",
                "Self funds",
                "Take over - if fully paid & Part paid meaing some of the amount being paid",
                "Pre-paid - Self funds",
                "Fore-closure - Take over",
                "",
            ]
        )
        for bank, cells, url in bank_rows:
            w.writerow([bank, *cells, url or ""])


def write_structured(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=MASTER_HEADERS)
        w.writeheader()
        for row in rows:
            w.writerow({h: row.get(h) if row.get(h) is not None else "" for h in MASTER_HEADERS})


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=None, help="Legacy CSV path")
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Home Loans.xlsx path")
    ap.add_argument("--sheet", type=str, default=DEFAULT_SHEET)
    ap.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    ap.add_argument(
        "--export-csv",
        type=Path,
        default=None,
        help="When reading xlsx, also write a synced CSV export",
    )
    args = ap.parse_args()

    if args.input is not None:
        if not args.input.exists():
            print(f"Missing input: {args.input}", file=sys.stderr)
            return 1
        rows = structure(args.input, from_xlsx=False)
        source_label = str(args.input)
    else:
        if not args.xlsx.exists():
            print(f"Missing xlsx: {args.xlsx}", file=sys.stderr)
            return 1
        rows = structure(args.xlsx, from_xlsx=True, sheet_name=args.sheet)
        source_label = f"{args.xlsx}!{args.sheet}"
        if args.export_csv:
            export_prepay_sheet_to_csv(args.xlsx, args.export_csv, sheet_name=args.sheet)
            print(f"Synced CSV export: {args.export_csv}")

    write_structured(args.output, rows)
    charged = sum(1 for r in rows if r["charge_status"] == "charged")
    nils = sum(1 for r in rows if r["charge_status"] == "nil")
    banks = sorted({r["bank_name"] for r in rows})
    print(f"Source: {source_label}")
    print(f"Wrote {args.output}")
    print(f"rows={len(rows)} charged={charged} nil={nils} banks_with_rows={len(banks)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Clean processing fees in HOME_LOANS_COMPARE_v1.xlsx: strip CIBIL, dedupe, Canara 0.25% removal."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "data" / "HOME_LOANS_COMPARE_v1.xlsx"

DEDUPE_KEY = [
    "bank_name", "scheme", "purpose", "facility_type", "rate_type", "occupation",
    "borrower_category", "loan_amount_band_applicable", "loan_amount_min", "loan_amount_max",
    "tenure_band_applicable", "tenure_months_min", "tenure_months_max",
    "charge_type", "percentage", "fixed_amount", "charge_min", "charge_max",
    "valid_from", "valid_till",
]


def blank(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def num(v):
    if blank(v):
        return None
    return float(v)


def row_key(d: dict) -> tuple:
    return tuple(d.get(f) for f in DEDUPE_KEY)


def proc_sort_key(d: dict) -> tuple:
    cid = str(d.get("charge_row_id") or "")
    try:
        n = int(cid.rsplit("-", 1)[-1])
    except ValueError:
        n = 0
    return (d.get("bank_name") or "", n)


def main() -> int:
    wb = load_workbook(MASTER)
    ws = wb["Bank_charges"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers)}

    all_rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        all_rows.append({h: raw[i] if i < len(raw) else None for h, i in hidx.items()})

    proc = [r for r in all_rows if r.get("origin") == "Offers.processing"]
    before = len(proc)
    stats = defaultdict(int)

    for r in proc:
        if r.get("cibil_band_applicable") != "No":
            stats["cibil_cleared"] += 1
        r["cibil_band_applicable"] = "No"
        r["cibil_band_score_min"] = None
        r["cibil_band_score_max"] = None

    # Canara: drop 0.25% tier (not in human Processing fees sheet)
    proc_filtered = []
    for r in proc:
        if r.get("bank_name") == "Canara Bank" and num(r.get("percentage")) == 0.0025:
            stats["canara_025_deleted"] += 1
            continue
        proc_filtered.append(r)
    proc = proc_filtered

    # Dedupe — keep earliest charge_row_id per key
    by_key: dict[tuple, dict] = {}
    for r in sorted(proc, key=proc_sort_key):
        k = row_key(r)
        if k not in by_key:
            by_key[k] = r
        else:
            stats["duplicates_deleted"] += 1

    keep_ids = {r["charge_row_id"] for r in by_key.values()}
    cleaned_by_id = {r["charge_row_id"]: r for r in by_key.values()}

    final: list[dict] = []
    for r in all_rows:
        if r.get("origin") != "Offers.processing":
            final.append(r)
            continue
        cid = r.get("charge_row_id")
        if cid in keep_ids:
            final.append(cleaned_by_id[cid])

    after = sum(1 for r in final if r.get("origin") == "Offers.processing")

    ws.delete_rows(2, ws.max_row)
    for rnum, row in enumerate(final, start=2):
        for col, h in enumerate(headers, start=1):
            ws.cell(row=rnum, column=col, value=row.get(h))

    wb.save(MASTER)
    wb.close()

    print(f"Processing rows: {before} → {after} (removed {before - after})")
    print(f"  CIBIL fields cleared: {stats['cibil_cleared']}")
    print(f"  Duplicate rows removed: {stats['duplicates_deleted']}")
    print(f"  Canara 0.25% rows removed: {stats['canara_025_deleted']}")
    print(f"Saved {MASTER}")

    by_bank = defaultdict(int)
    for r in final:
        if r.get("origin") == "Offers.processing":
            by_bank[r["bank_name"]] += 1
    print("\nPer bank:")
    for b in sorted(by_bank):
        print(f"  {b}: {by_bank[b]}")
    print(f"  TOTAL: {after}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
